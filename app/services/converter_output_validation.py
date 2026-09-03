# -*- coding: utf-8 -*-
"""Validação somente-leitura das saídas temporárias do /converter."""
from __future__ import annotations

import csv
import io
import json
import os
import re
import zipfile
from xml.etree import ElementTree
from typing import Callable, Optional


class ConverterOutputValidationError(RuntimeError):
    """A saída não corresponde estruturalmente ao formato solicitado."""


_OOXML_CONTENT_TYPES = "[Content_Types].xml"
_XLSM_MAIN_CONTENT_TYPE = (
    b"application/vnd.ms-excel.sheet.macroenabled.main+xml"
)
_SUPPORTED_TARGETS = {"pdf", "docx", "xlsx", "xlsm", "csv"}
_CSV_KNOWN_DIAGNOSTICS = frozenset({
    "Falha ao extrair conteúdo do PDF.",
})
_CSV_PARTIAL_INVALID_MARKER = "GVCSV_PARTIAL_INVALID"
_CSV_ERROR_JSON_KEYS = frozenset({"error", "errors", "traceback"})
_CSV_HTML_PREFIX_RE = re.compile(
    r"^(?:<!doctype\s+html\b|<html\b)",
    flags=re.IGNORECASE,
)


def _checkpoint(check_deadline: Optional[Callable[[], None]]) -> None:
    if check_deadline is not None:
        check_deadline()


def _validate_zip_members(path: str, required: set[str]) -> bytes:
    try:
        with zipfile.ZipFile(path, "r") as archive:
            if archive.testzip() is not None:
                raise ConverterOutputValidationError("Pacote OOXML corrompido.")
            names = set(archive.namelist())
            if not required.issubset(names):
                raise ConverterOutputValidationError(
                    "Estrutura OOXML mínima ausente."
                )
            content_types = archive.read(_OOXML_CONTENT_TYPES)
            ElementTree.fromstring(content_types)
            return content_types
    except ConverterOutputValidationError:
        raise
    except Exception as exc:
        raise ConverterOutputValidationError(
            "Pacote OOXML inválido."
        ) from exc


def _validate_pdf(path: str) -> None:
    try:
        import pikepdf

        with pikepdf.open(path) as document:
            if len(document.pages) < 1:
                raise ConverterOutputValidationError("PDF sem páginas.")
    except ConverterOutputValidationError:
        raise
    except Exception as exc:
        raise ConverterOutputValidationError("PDF ilegível.") from exc


def _validate_docx(path: str) -> None:
    _validate_zip_members(
        path,
        {_OOXML_CONTENT_TYPES, "word/document.xml"},
    )
    try:
        with zipfile.ZipFile(path, "r") as archive:
            ElementTree.fromstring(archive.read("word/document.xml"))
    except Exception as exc:
        raise ConverterOutputValidationError(
            "Documento OOXML inválido."
        ) from exc


def _validate_workbook(path: str, *, keep_vba: bool) -> None:
    required = {_OOXML_CONTENT_TYPES, "xl/workbook.xml"}
    content_types = _validate_zip_members(path, required)
    if keep_vba and _XLSM_MAIN_CONTENT_TYPE not in content_types.lower():
        raise ConverterOutputValidationError(
            "A saída XLSM não possui content type macro-enabled."
        )

    workbook = None
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(
            path,
            read_only=True,
            data_only=False,
            keep_vba=keep_vba,
        )
        has_useful_cell = False
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                if any(
                    value is not None and str(value).strip()
                    for value in row
                ):
                    has_useful_cell = True
                    break
            if has_useful_cell:
                break
        if not has_useful_cell:
            raise ConverterOutputValidationError(
                "Workbook OOXML sem conteúdo útil."
            )
    except ConverterOutputValidationError:
        raise
    except Exception as exc:
        raise ConverterOutputValidationError(
            "Workbook OOXML ilegível."
        ) from exc
    finally:
        if workbook is not None:
            workbook.close()
            vba_archive = getattr(workbook, "vba_archive", None)
            if vba_archive is not None:
                vba_archive.close()


def _csv_contains_binary_control(payload: bytes) -> bool:
    return any(
        (byte < 32 and byte not in {9, 10, 13}) or byte == 127
        for byte in payload
    )


def _csv_is_known_error_document(text: str) -> bool:
    stripped = text.strip()
    if stripped in _CSV_KNOWN_DIAGNOSTICS:
        return True
    first_line = stripped.splitlines()[0].strip() if stripped else ""
    if first_line == _CSV_PARTIAL_INVALID_MARKER:
        return True
    if _CSV_HTML_PREFIX_RE.match(stripped) is not None:
        return True
    if stripped.startswith("Traceback (most recent call last):"):
        return True
    try:
        document = json.loads(stripped)
    except (TypeError, ValueError, json.JSONDecodeError):
        return False
    return (
        isinstance(document, dict)
        and bool(_CSV_ERROR_JSON_KEYS.intersection(document))
    )


def _csv_reader(text: str):
    sample = text[: 64 * 1024]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;")
        return csv.reader(
            io.StringIO(text, newline=""),
            dialect=dialect,
            strict=True,
        )
    except csv.Error:
        delimiter = ";" if sample.count(";") > sample.count(",") else ","
        return csv.reader(
            io.StringIO(text, newline=""),
            delimiter=delimiter,
            strict=True,
        )


def _validate_csv(path: str, *, require_table_data: bool = False) -> None:
    try:
        with open(path, "rb") as fh:
            payload = fh.read()
        if b"\x00" in payload:
            raise ConverterOutputValidationError("CSV contém byte NUL.")
        if _csv_contains_binary_control(payload):
            raise ConverterOutputValidationError("CSV contém conteúdo binário.")
        text = payload.decode("utf-8-sig", errors="strict")
    except ConverterOutputValidationError:
        raise
    except (OSError, UnicodeError) as exc:
        raise ConverterOutputValidationError("CSV inválido.") from exc

    if not text.strip():
        raise ConverterOutputValidationError("CSV sem conteúdo útil.")
    if _csv_is_known_error_document(text):
        raise ConverterOutputValidationError("CSV contém diagnóstico interno.")

    expected_columns = None
    meaningful_rows = 0
    try:
        for row in _csv_reader(text):
            if not row or not any(cell.strip() for cell in row):
                continue
            if expected_columns is None:
                expected_columns = len(row)
            elif len(row) != expected_columns:
                raise ConverterOutputValidationError(
                    "CSV contém quantidade inconsistente de colunas."
                )
            meaningful_rows += 1
    except ConverterOutputValidationError:
        raise
    except csv.Error as exc:
        raise ConverterOutputValidationError("CSV estruturalmente inválido.") from exc

    if meaningful_rows == 0:
        raise ConverterOutputValidationError("CSV sem conteúdo útil.")
    if require_table_data and meaningful_rows < 2:
        raise ConverterOutputValidationError(
            "CSV de tabela não contém linha de dados."
        )


def validate_converter_output(
    path: str,
    workdir: str,
    target: str,
    *,
    check_deadline: Optional[Callable[[], None]] = None,
    source_ext: Optional[str] = None,
    require_table_data: bool = False,
) -> str:
    """Valida uma saída ainda não publicada e devolve seu caminho real."""
    normalized_target = (target or "").strip().lower()
    if normalized_target not in _SUPPORTED_TARGETS:
        raise ConverterOutputValidationError("Target de validação inválido.")

    workdir_abs = os.path.abspath(workdir)
    candidate_abs = os.path.abspath(path or "")
    workdir_real = os.path.realpath(workdir_abs)
    candidate_real = os.path.realpath(candidate_abs)
    try:
        contained = (
            candidate_real != workdir_real
            and os.path.commonpath([workdir_real, candidate_real])
            == workdir_real
        )
    except ValueError:
        contained = False

    if (
        not contained
        or os.path.islink(candidate_abs)
        or not os.path.isfile(candidate_abs)
    ):
        raise ConverterOutputValidationError(
            "Saída temporária fora do diretório isolado."
        )
    try:
        if os.path.getsize(candidate_abs) <= 0:
            raise ConverterOutputValidationError("Saída temporária vazia.")
    except OSError as exc:
        raise ConverterOutputValidationError(
            "Saída temporária ilegível."
        ) from exc

    extension = os.path.splitext(candidate_abs)[1].lower()
    if extension != f".{normalized_target}":
        raise ConverterOutputValidationError(
            "Extensão da saída não corresponde ao target."
        )

    _checkpoint(check_deadline)
    if normalized_target == "pdf":
        _validate_pdf(candidate_abs)
    elif normalized_target == "docx":
        _validate_docx(candidate_abs)
    elif normalized_target == "xlsx":
        _validate_workbook(candidate_abs, keep_vba=False)
    elif normalized_target == "xlsm":
        _validate_workbook(candidate_abs, keep_vba=True)
    else:
        normalized_source_ext = (source_ext or "").strip().lower().lstrip(".")
        _validate_csv(
            candidate_abs,
            require_table_data=(
                require_table_data or normalized_source_ext == "pdf"
            ),
        )
    _checkpoint(check_deadline)
    return candidate_real

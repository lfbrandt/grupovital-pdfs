# -*- coding: utf-8 -*-
"""
converter_service.py — Conversões e utilidades PDF/planilhas
Principais pontos:
- Imagem → PDF sai em A4 (ou Letter), centralizada, com margens mínimas. Auto-paisagem opcional.
- Respeita EXIF Orientation (ImageOps.exif_transpose) antes de qualquer cálculo de layout.
- Merge: pode normalizar para A4/Letter com PDFFitPage e AutoRotate configurável (none/page/all).
- PDF → XLSX no estilo “modelo” (já existente), com OCR opcional e vários fallbacks.
"""
from __future__ import annotations

import contextvars
import csv
import importlib.util
import io
import json
import math
import os, re, tempfile, subprocess, shutil, logging, time, platform
import secrets
import sys
import unicodedata
import zipfile
from collections import Counter
from contextlib import contextmanager
from dataclasses import dataclass, replace
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterator, List, Optional, Dict, Any, Tuple, Sequence
from xml.etree import ElementTree

from PIL import Image, ImageOps
from werkzeug.exceptions import BadRequest

from ..utils.limits import enforce_pdf_page_limit
# 🔒 sandbox (mesmo mecanismo usado no merge_service)
from .sandbox import run_in_sandbox
from . import camelot_worker as _camelot_worker
from .pdf_table_geometry import (
    CandidateGeometryReport,
    PageTableGeometry,
    geometry_report_for_candidate,
    map_pdf_table_geometry,
)
from flask import has_request_context, request  # (para tentar bytes_in via Content-Length)
from ..utils.stats import record_job_event      # (7.1) métricas

logger = logging.getLogger(__name__)

IMG_EXTS   = {'jpg','jpeg','png','bmp','tif','tiff','webp'}
DOC_EXTS   = {'doc','docx','odt','rtf','txt','html','htm','ppt','pptx','odp'}
SHEET_EXTS = {'csv','xls','xlsx','ods'}


class ConverterTimeoutError(RuntimeError):
    """O prazo total ou o timeout de uma ferramenta foi excedido."""


class _ConverterDeadlineReservedError(ConverterTimeoutError):
    """A margem final do deadline deve ser preservada para cleanup/resposta."""


class ConverterToolUnavailableError(RuntimeError):
    """Uma ferramenta externa obrigatória não está disponível."""


class ConverterToolExecutionError(RuntimeError):
    """Uma ferramenta externa terminou sem produzir uma saída utilizável."""


class ConverterNoTableError(RuntimeError):
    """O PDF foi processado, mas não contém uma tabela útil."""


class ConverterExtractionError(RuntimeError):
    """Os extratores não conseguiram produzir uma tabela segura."""


class _ConverterJobRuntime:
    def __init__(self, seconds: float):
        try:
            duration = float(seconds)
        except (TypeError, ValueError):
            duration = 300.0
        if duration <= 0:
            duration = 300.0
        self.deadline = time.monotonic() + duration
        self._temporary_paths: List[str] = []

    def remaining(self, stage: str = "job") -> float:
        remaining = self.deadline - time.monotonic()
        if remaining <= 0:
            raise ConverterTimeoutError(
                f"Prazo do converter esgotado em {stage}."
            )
        return remaining

    def effective_timeout(self, tool_timeout: float, stage: str) -> float:
        try:
            configured = float(tool_timeout)
        except (TypeError, ValueError):
            configured = 1.0
        if configured <= 0:
            configured = 1.0
        return min(configured, self.remaining(stage))

    def register_temporary(self, path: str) -> None:
        self._temporary_paths.append(os.path.abspath(path))

    def cleanup(self) -> None:
        for path in reversed(self._temporary_paths):
            try:
                if os.path.isfile(path) or os.path.islink(path):
                    os.remove(path)
            except OSError:
                pass
        self._temporary_paths.clear()


_ACTIVE_CONVERTER_RUNTIME: contextvars.ContextVar[
    Optional[_ConverterJobRuntime]
] = contextvars.ContextVar("converter_job_runtime", default=None)


@contextmanager
def converter_job_runtime(seconds: float) -> Iterator[_ConverterJobRuntime]:
    """Cria o único deadline monotônico do job atual do converter."""
    runtime = _ConverterJobRuntime(seconds)
    token = _ACTIVE_CONVERTER_RUNTIME.set(runtime)
    try:
        yield runtime
    finally:
        runtime.cleanup()
        _ACTIVE_CONVERTER_RUNTIME.reset(token)


def check_converter_deadline(stage: str = "job") -> None:
    runtime = _ACTIVE_CONVERTER_RUNTIME.get()
    if runtime is not None:
        runtime.remaining(stage)


def _effective_converter_timeout(tool_timeout: float, stage: str) -> float:
    runtime = _ACTIVE_CONVERTER_RUNTIME.get()
    if runtime is None:
        try:
            value = float(tool_timeout)
        except (TypeError, ValueError):
            value = 1.0
        return value if value > 0 else 1.0
    return runtime.effective_timeout(tool_timeout, stage)


def _register_converter_temporary(path: str) -> None:
    runtime = _ACTIVE_CONVERTER_RUNTIME.get()
    if runtime is not None:
        runtime.register_temporary(path)

# =========================
# Ghostscript configuration
# =========================
env_gs = os.environ.get("GS_BIN") or os.environ.get("GHOSTSCRIPT_BIN")
if env_gs:
    GHOSTSCRIPT_BIN = env_gs
elif platform.system() == "Windows":
    GHOSTSCRIPT_BIN = "gswin64c"
else:
    GHOSTSCRIPT_BIN = "gs"

_GS_TO = os.environ.get("GS_TIMEOUT") or os.environ.get("GHOSTSCRIPT_TIMEOUT") or "60"
GHOSTSCRIPT_TIMEOUT = int(_GS_TO)

# =========================
# Merge normalization envs
# =========================
MERGE_NORMALIZE_MODE = (os.environ.get("MERGE_NORMALIZE_MODE", "auto") or "auto").lower()  # auto|always|off
MERGE_NORMALIZE_AUTOROTATE = (os.environ.get("MERGE_NORMALIZE_AUTOROTATE", "none") or "none").lower()  # none|page|all
MERGE_STRIP_ROTATE = (os.environ.get("MERGE_STRIP_ROTATE", "0") == "1")

# Tamanhos padrão em pontos (1/72")
SIZES_PT = {
    "A4":     (595.2756, 841.8898),  # 210 x 297 mm
    "LETTER": (612.0,   792.0),      # 8.5 x 11 in
}

# ---------------- TMP helpers ----------------
def _save_upload_to_tmp(
    upload_file,
    suffix: str,
    *,
    directory: Optional[str] = None,
) -> str:
    fd, path = tempfile.mkstemp(suffix=suffix, dir=directory)
    os.close(fd)
    upload_file.stream.seek(0)
    with open(path, 'wb') as out:
        shutil.copyfileobj(upload_file.stream, out)
    return path

def _tmp_out_path(ext: str) -> str:
    fd, path = tempfile.mkstemp(suffix='.' + ext.lstrip('.')); os.close(fd)
    return path

def _unique_out_path(out_dir: str, base: str, ext: str) -> str:
    out_dir = os.path.abspath(out_dir); os.makedirs(out_dir, exist_ok=True)
    candidate = os.path.join(out_dir, f"{base}.{ext}"); i = 1
    while os.path.exists(candidate):
        candidate = os.path.join(out_dir, f"{base} ({i}).{ext}"); i += 1
    return candidate

# ======================================================================
# Imagem → PDF A4/Letter (ReportLab; fallback via PIL mantendo página real)
# Env:
#   IMG2PDF_MODE=fit|cover (default fit)   → fit mantém tudo visível; cover “preenche” (pode cortar)
#   IMG2PDF_MARGIN_PT=18                   → margem em pontos
#   IMG2PDF_LANDSCAPE_AUTO=1|0 (default 1) → paisagem automática se imagem deitada
#   IMG2PDF_DPI=300                        → usado no fallback PIL
#   IMG2PDF_PAGE_SIZE=A4|LETTER            → força tamanho; default A4
# ======================================================================
def _apply_exif(img: Image.Image) -> Image.Image:
    """Aplica rotação EXIF (se houver) antes de qualquer conversão/resize."""
    try:
        return ImageOps.exif_transpose(img)
    except Exception:
        return img

def _image_to_pdf(in_path: str, out_path: str) -> None:
    mode = (os.environ.get("IMG2PDF_MODE", "fit") or "fit").lower()
    margin = float(os.environ.get("IMG2PDF_MARGIN_PT", "18"))
    auto_land = os.environ.get("IMG2PDF_LANDSCAPE_AUTO", "1") == "1"
    page_size_name = (os.environ.get("IMG2PDF_PAGE_SIZE", "A4") or "A4").upper()
    base_w_pt, base_h_pt = SIZES_PT.get(page_size_name, SIZES_PT["A4"])

    # Tenta ReportLab (precisão A4/Letter garantida)
    try:
        from reportlab.pdfgen import canvas as _rl_canvas
        from reportlab.lib.pagesizes import A4 as _A4, LETTER as _LETTER, landscape as _landscape
        from reportlab.lib.utils import ImageReader

        size_map = {"A4": _A4, "LETTER": _LETTER}
        base_size = size_map.get(page_size_name, _A4)

        im = Image.open(in_path)
        frames: List[Image.Image] = []
        try:
            # coleta frames (TIFF multipáginas etc.)
            i = 0
            while True:
                im.seek(i)
                fr = _apply_exif(im.copy())
                frames.append(fr.convert("RGB"))
                i += 1
        except EOFError:
            pass
        if not frames:
            frames = [_apply_exif(im.copy()).convert("RGB")]

        c: Optional[_rl_canvas.Canvas] = None
        for f in frames:
            iw, ih = f.size
            pagesize = base_size
            if auto_land and iw > ih * 1.05:
                pagesize = _landscape(base_size)
            PW, PH = pagesize

            max_w, max_h = PW - 2*margin, PH - 2*margin
            scale = max(max_w / iw, max_h / ih) if mode == "cover" else min(max_w / iw, max_h / ih)
            tw, th = iw * scale, ih * scale
            x, y = (PW - tw) / 2.0, (PH - th) / 2.0

            # trata transparência para fundo branco
            if f.mode in ("RGBA", "LA", "P"):
                bg = Image.new("RGB", f.size, (255, 255, 255))
                if f.mode in ("RGBA", "LA"):
                    bg.paste(f, mask=f.split()[-1])
                else:
                    bg.paste(f)
                f = bg

            if c is None:
                c = _rl_canvas.Canvas(out_path, pagesize=pagesize)
            else:
                c.setPageSize(pagesize)

            c.drawImage(ImageReader(f), x, y, width=tw, height=th,
                        preserveAspectRatio=True, mask='auto')
            c.showPage()
        if c:
            c.save()
        try:
            im.close()
        except Exception:
            pass
        return
    except Exception as e:
        logger.debug("ReportLab indisponível (%s). Usando fallback PIL para A4/Letter.", e)

    # Fallback PIL: cria uma “lona” A4/Letter em pixels (DPI alto) e salva como PDF
    dpi = int(os.environ.get("IMG2PDF_DPI", "300"))
    base_w_px = int(round(base_w_pt / 72.0 * dpi))
    base_h_px = int(round(base_h_pt / 72.0 * dpi))
    margin_px = int(round(margin / 72.0 * dpi))

    im = Image.open(in_path)
    pages: List[Image.Image] = []
    try:
        frames: List[Image.Image] = []
        try:
            i = 0
            while True:
                im.seek(i)
                frm = _apply_exif(im.copy()).convert("RGB")
                frames.append(frm)
                i += 1
        except EOFError:
            pass
        if not frames:
            frames = [_apply_exif(im.copy()).convert("RGB")]

        for f in frames:
            # auto paisagem
            W, H = (base_w_px, base_h_px)
            if auto_land and f.width > f.height * 1.05:
                W, H = (base_h_px, base_w_px)

            canvas_img = Image.new("RGB", (W, H), (255, 255, 255))
            max_w, max_h = W - 2*margin_px, H - 2*margin_px

            scale = max(max_w / f.width, max_h / f.height) if mode == "cover" else min(max_w / f.width, max_h / f.height)
            tw, th = max(1, int(round(f.width * scale))), max(1, int(round(f.height * scale)))
            f2 = f.resize((tw, th), Image.LANCZOS)
            x, y = (W - tw) // 2, (H - th) // 2
            canvas_img.paste(f2, (x, y))
            pages.append(canvas_img)

        pages[0].save(out_path, save_all=True, append_images=pages[1:],
                      format='PDF', resolution=float(dpi))
    finally:
        try: im.close()
        except Exception: pass

# ---------------- LibreOffice helpers ----------------
# Caminhos padrão do LibreOffice no Windows (adicione mais se necessário)
_LO_WIN_PATHS = [
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    r"C:\Program Files\LibreOffice 7\program\soffice.exe",
    r"C:\Program Files\LibreOffice 6\program\soffice.exe",
]

def _soffice_bin() -> str:
    """
    Resolve o binário do soffice/LibreOffice.
    Ordem de preferência:
      1. SOFFICE_BIN (env)
      2. LIBREOFFICE_BIN (env)
      3. shutil.which("soffice") / shutil.which("soffice.com") / shutil.which("libreoffice")
      4. Caminhos fixos comuns no Windows
    Levanta erro controlado se não encontrar.
    """
    import shutil as _sh

    # 1. Variáveis de ambiente explícitas
    from_env = os.environ.get('SOFFICE_BIN') or os.environ.get('LIBREOFFICE_BIN')
    if from_env:
        if os.path.isfile(from_env):
            logger.info("[converter] LibreOffice resolvido via configuração")
            return from_env
        logger.warning("[converter] configuração do LibreOffice não foi resolvida")

    # 2. PATH do sistema
    for candidate in ("soffice", "soffice.com", "libreoffice"):
        found = _sh.which(candidate)
        if found:
            logger.info("[converter] LibreOffice resolvido via PATH")
            return found

    # 3. Caminhos fixos Windows
    if os.name == 'nt':
        for path in _LO_WIN_PATHS:
            if os.path.isfile(path):
                logger.info("[converter] LibreOffice resolvido no Windows")
                return path

    raise ConverterToolUnavailableError(
        "LibreOffice não está disponível."
    )


def _positive_env_int(names: Tuple[str, ...], default: int) -> int:
    for name in names:
        raw = os.environ.get(name)
        if raw is None or not str(raw).strip():
            continue
        try:
            value = int(str(raw).strip())
        except (TypeError, ValueError):
            continue
        return value if value > 0 else default
    return default


def _lo_convert(in_path: str, out_dir: str, out_ext: str,
                filter_name: Optional[str] = None,
                filter_opts: Optional[str] = None,
                input_filter: Optional[str] = None) -> str:
    convert_to = out_ext
    if filter_name and filter_opts:
        convert_to = f"{out_ext}:{filter_name}:{filter_opts}"
    elif filter_name:
        convert_to = f"{out_ext}:{filter_name}"

    lo_timeout = _positive_env_int(
        ("LO_TIMEOUT", "LO_CONVERT_TIMEOUT_SEC", "LIBREOFFICE_TIMEOUT"),
        120,
    )
    effective_timeout = _effective_converter_timeout(
        lo_timeout,
        "libreoffice",
    )
    soffice = _soffice_bin()
    out_dir = os.path.abspath(out_dir)
    os.makedirs(out_dir, exist_ok=True)
    profile_dir = tempfile.mkdtemp(prefix=".lo-profile-", dir=out_dir)
    profile_uri = Path(profile_dir).resolve().as_uri()
    input_filter_args = (
        [f'--infilter={input_filter}']
        if input_filter
        else []
    )

    cmd = [
        soffice,
        f"-env:UserInstallation={profile_uri}",
        '--headless', '--safe-mode', '--nologo', '--nodefault', '--nolockcheck', '--invisible',
        *input_filter_args,
        '--convert-to', convert_to, '--outdir', out_dir, in_path
    ]
    try:
        proc = run_in_sandbox(
            cmd,
            cwd=out_dir,
            timeout=effective_timeout,
            cpu_seconds=max(1, math.ceil(effective_timeout)),
            mem_mb=1024,
            output_limit_chars=4096,
        )
    except FileNotFoundError:
        raise ConverterToolUnavailableError(
            "LibreOffice não está disponível."
        ) from None
    except subprocess.TimeoutExpired:
        raise ConverterTimeoutError(
            "LibreOffice excedeu o tempo permitido."
        ) from None
    finally:
        shutil.rmtree(profile_dir, ignore_errors=True)

    if getattr(proc, "returncode", 0) != 0:
        raise ConverterToolExecutionError(
            "LibreOffice terminou com erro."
        )

    base = os.path.splitext(os.path.basename(in_path))[0]
    produced = os.path.join(out_dir, f"{base}.{out_ext}")
    if not os.path.exists(produced):
        for fn in os.listdir(out_dir):
            if fn.lower().startswith(base.lower()+".") and fn.lower().endswith("."+out_ext):
                produced = os.path.join(out_dir, fn); break
    if not os.path.exists(produced):
        raise ConverterToolExecutionError(
            "LibreOffice não gerou a saída esperada."
        )
    check_converter_deadline("libreoffice-output")
    return produced


def libreoffice_healthcheck(timeout: float = 5) -> str:
    """Consulta a versão via executor sem expor caminhos ou stderr."""
    soffice = _soffice_bin()
    effective_timeout = max(0.1, min(float(timeout), 5.0))
    with tempfile.TemporaryDirectory(prefix="gvpdf_lo_health_") as workdir:
        profile_dir = tempfile.mkdtemp(prefix=".lo-profile-", dir=workdir)
        try:
            cmd = [
                soffice,
                f"-env:UserInstallation={Path(profile_dir).resolve().as_uri()}",
                "--headless",
                "--version",
            ]
            result = run_in_sandbox(
                cmd,
                cwd=workdir,
                timeout=effective_timeout,
                cpu_seconds=max(1, math.ceil(effective_timeout)),
                mem_mb=512,
                output_limit_chars=512,
            )
        except FileNotFoundError:
            raise ConverterToolUnavailableError(
                "LibreOffice não está disponível."
            ) from None
        except subprocess.TimeoutExpired:
            raise ConverterTimeoutError(
                "Healthcheck do LibreOffice excedeu o tempo permitido."
            ) from None
        finally:
            shutil.rmtree(profile_dir, ignore_errors=True)
    if result.returncode != 0:
        raise ConverterToolExecutionError(
            "Falha no healthcheck do LibreOffice."
        )
    first_line = (result.stdout or "").strip().splitlines()
    candidate = first_line[0] if first_line else ""
    version_match = re.search(
        r"\bLibreOffice\s+[0-9][0-9A-Za-z.+-]*",
        candidate,
        flags=re.IGNORECASE,
    )
    return (
        version_match.group(0)
        if version_match is not None
        else "LibreOffice disponível"
    )

# ---------------- Camelot / Ghostscript env ----------------
def _prepare_camelot_env() -> None:
    import shutil as _sh
    gs = (os.environ.get("GS_BIN") or os.environ.get("GHOSTSCRIPT_BIN")
          or _sh.which("gswin64c") or _sh.which("gs"))
    if not gs:
        raise ConverterToolUnavailableError(
            "Ghostscript não está disponível."
        )
    os.environ["PATH"] = os.path.dirname(gs) + os.pathsep + os.environ.get("PATH", "")
    os.environ["GHOSTSCRIPT_PATH"] = gs
    os.environ["GS_PROG"] = gs
    try:
        import cv2  # noqa: F401
    except Exception as e:
        raise ConverterToolUnavailableError(
            "OpenCV não está disponível."
        ) from e

# ---------------- OCR helper ----------------
def _pdf_has_selectable_text(in_pdf: str) -> bool:
    try:
        import pdfplumber
        with pdfplumber.open(in_pdf) as pdf:
            for page in pdf.pages:
                if (page.chars or []) or (page.extract_text() or "").strip():
                    return True
    except Exception:
        pass
    return False

def _try_ocr(in_pdf: str) -> str:
    if os.environ.get("OCR_ON_PDF_TO_XLSX","0") != "1":
        return in_pdf
    executable = shutil.which("ocrmypdf")
    if executable:
        cmd_prefix = [executable]
    elif importlib.util.find_spec("ocrmypdf") is not None:
        cmd_prefix = [sys.executable, "-m", "ocrmypdf"]
    else:
        raise ConverterToolUnavailableError(
            "OCRmyPDF não está disponível."
        )
    enforce_pdf_page_limit(in_pdf, label="PDF para OCR")
    output_dir = os.path.dirname(os.path.abspath(in_pdf))
    fd, out_pdf = tempfile.mkstemp(
        prefix=".ocr-",
        suffix=".pdf",
        dir=output_dir,
    )
    os.close(fd)
    _register_converter_temporary(out_pdf)
    ocr_lang = os.environ.get("OCR_LANGS","por+eng")
    ocr_timeout = _positive_env_int(
        ("OCR_TIMEOUT", "OCR_TIMEOUT_SEC", "TESSERACT_TIMEOUT"),
        120,
    )
    ocr_jobs = min(
        4,
        _positive_env_int(("OCR_JOBS",), 1),
    )
    effective_timeout = _effective_converter_timeout(
        ocr_timeout,
        "ocr",
    )
    cmd = cmd_prefix + ["--skip-text","--force-ocr","--rotate-pages",
           "--jobs",str(ocr_jobs),
           "--tesseract-timeout","60","-l",ocr_lang,in_pdf,out_pdf]
    try:
        result = run_in_sandbox(
            cmd,
            cwd=output_dir,
            timeout=effective_timeout,
            cpu_seconds=max(1, math.ceil(effective_timeout)),
            mem_mb=1024,
            output_limit_chars=4096,
        )
    except FileNotFoundError:
        raise ConverterToolUnavailableError(
            "OCRmyPDF não está disponível."
        ) from None
    except subprocess.TimeoutExpired:
        raise ConverterTimeoutError(
            "OCR excedeu o tempo permitido."
        ) from None
    if result.returncode != 0:
        raise ConverterToolExecutionError("OCR terminou com erro.")
    try:
        import pikepdf

        with pikepdf.open(out_pdf) as document:
            if len(document.pages) < 1:
                raise ConverterToolExecutionError(
                    "OCR gerou PDF sem páginas."
                )
    except ConverterToolExecutionError:
        raise
    except Exception as exc:
        raise ConverterToolExecutionError(
            "OCR gerou uma saída inválida."
        ) from exc
    check_converter_deadline("ocr-output")
    return out_pdf

# ---------------- Excel helpers (mantidos) ----------------
EXCEL_DANGEROUS_PREFIXES = ("=", "+", "-", "@")


def _neutralize_spreadsheet_formula(value: Any) -> Any:
    """Mantém tipos e força texto não confiável a permanecer literal no Excel."""
    if not isinstance(value, str) or value == "":
        return value

    position = 0
    leading_control = False
    while position < len(value):
        character = value[position]
        category = unicodedata.category(character)
        if not (character.isspace() or category.startswith("C")):
            break
        if category.startswith("C"):
            leading_control = True
        position += 1

    first_significant = value[position] if position < len(value) else ""
    if leading_control or first_significant in EXCEL_DANGEROUS_PREFIXES:
        return "'" + value
    return value


_CSV_UNAMBIGUOUS_NUMBER_RE = re.compile(
    r'^[+-]?(?:[0-9]+(?:[.,][0-9]+)?|[.,][0-9]+)'
    r'(?:[eE][+-]?[0-9]+)?$'
)


def _csv_significant_text(value: str) -> str:
    start = 0
    end = len(value)
    while start < end:
        character = value[start]
        if not (
            character.isspace()
            or unicodedata.category(character).startswith('C')
        ):
            break
        start += 1
    while end > start:
        character = value[end - 1]
        if not (
            character.isspace()
            or unicodedata.category(character).startswith('C')
        ):
            break
        end -= 1
    return value[start:end]


def _neutralize_csv_field_for_spreadsheet(value: Any) -> Any:
    '''Protege texto CSV, preservando apenas tokens numéricos completos.'''
    if not isinstance(value, str) or value == '':
        return value
    significant = _csv_significant_text(value)
    if (
        significant
        and _CSV_UNAMBIGUOUS_NUMBER_RE.fullmatch(significant) is not None
    ):
        return value
    return _neutralize_spreadsheet_formula(value)


def _csv_delimiter_outside_quotes(text: str) -> str:
    counts = {',': 0, ';': 0}
    in_quotes = False
    position = 0
    while position < len(text):
        character = text[position]
        if character == chr(34):
            if (
                in_quotes
                and position + 1 < len(text)
                and text[position + 1] == chr(34)
            ):
                position += 2
                continue
            in_quotes = not in_quotes
        elif not in_quotes and character in counts:
            counts[character] += 1
        position += 1
    return ';' if counts[';'] > counts[','] else ','


def _csv_dialect_parameters(text: str) -> Dict[str, Any]:
    sample = text[:65536]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;')
    except csv.Error:
        return {
            'delimiter': _csv_delimiter_outside_quotes(text),
            'quotechar': chr(34),
            'doublequote': True,
            'escapechar': None,
            'skipinitialspace': False,
            'quoting': csv.QUOTE_MINIMAL,
        }
    return {
        'delimiter': dialect.delimiter,
        'quotechar': dialect.quotechar or chr(34),
        'doublequote': dialect.doublequote,
        'escapechar': dialect.escapechar,
        'skipinitialspace': False,
        'quoting': dialect.quoting,
    }


def _decode_uploaded_csv(raw: bytes) -> str:
    if b'\x00' in raw:
        raise BadRequest('CSV inválido para conversão.')
    for encoding in ('utf-8-sig', 'cp1252'):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise BadRequest('CSV inválido para conversão.')


def _write_neutralized_csv_copy(
    in_path: str,
    out_dir: str,
) -> Tuple[str, str]:
    '''Analisa o CSV e cria uma cópia temporária protegida para o Calc.'''
    protected_path = ''
    try:
        with open(in_path, 'rb') as source:
            text = _decode_uploaded_csv(source.read())
        parameters = _csv_dialect_parameters(text)
        reader = csv.reader(
            io.StringIO(text, newline=''),
            delimiter=parameters['delimiter'],
            quotechar=parameters['quotechar'],
            doublequote=parameters['doublequote'],
            escapechar=parameters['escapechar'],
            skipinitialspace=parameters['skipinitialspace'],
            strict=True,
        )
        rows = [
            [_neutralize_csv_field_for_spreadsheet(field) for field in row]
            for row in reader
        ]
        if '\r\n' in text:
            line_terminator = '\r\n'
        elif '\r' in text and '\n' not in text:
            line_terminator = '\r'
        else:
            line_terminator = '\n'

        descriptor, protected_path = tempfile.mkstemp(
            prefix='.csv-protected-',
            suffix='.csv',
            dir=out_dir,
        )
        with os.fdopen(
            descriptor,
            'w',
            encoding='utf-8-sig',
            newline='',
        ) as destination:
            writer = csv.writer(
                destination,
                delimiter=parameters['delimiter'],
                quotechar=parameters['quotechar'],
                doublequote=parameters['doublequote'],
                escapechar=parameters['escapechar'],
                quoting=csv.QUOTE_ALL,
                lineterminator=line_terminator,
            )
            writer.writerows(rows)
        return protected_path, parameters['delimiter']
    except BadRequest:
        if protected_path:
            try:
                os.remove(protected_path)
            except OSError:
                pass
        raise
    except (OSError, csv.Error, UnicodeError) as exc:
        if protected_path:
            try:
                os.remove(protected_path)
            except OSError:
                pass
        raise BadRequest('CSV inválido para conversão.') from exc


def _assert_spreadsheet_has_no_formulas(output_path: str) -> None:
    '''Rejeita saída OOXML que ainda contenha fórmula em uma worksheet.'''
    try:
        with zipfile.ZipFile(output_path, 'r') as archive:
            worksheet_names = [
                name
                for name in archive.namelist()
                if name.startswith('xl/worksheets/')
                and name.endswith('.xml')
            ]
            if not worksheet_names:
                raise ConverterToolExecutionError(
                    'A planilha convertida não contém worksheets válidas.'
                )
            for worksheet_name in worksheet_names:
                formula_found = False
                with archive.open(worksheet_name, 'r') as worksheet:
                    for _event, element in ElementTree.iterparse(
                        worksheet,
                        events=('start',),
                    ):
                        if element.tag.rsplit('}', 1)[-1] == 'f':
                            formula_found = True
                        element.clear()
                if formula_found:
                    raise ConverterToolExecutionError(
                        'A planilha convertida contém fórmula residual.'
                    )
    except ConverterToolExecutionError:
        raise
    except (
        OSError,
        KeyError,
        RuntimeError,
        zipfile.BadZipFile,
        ElementTree.ParseError,
    ) as exc:
        raise ConverterToolExecutionError(
            'Não foi possível verificar a segurança da planilha convertida.'
        ) from exc


def _convert_csv_to_spreadsheet(
    in_path: str,
    out_dir: str,
    out_ext: str,
    filter_name: str,
) -> str:
    protected_path, delimiter = _write_neutralized_csv_copy(in_path, out_dir)
    produced_path = ''
    expected_path = os.path.join(
        out_dir,
        f'{Path(protected_path).stem}.{out_ext}',
    )
    try:
        produced_path = _lo_convert(
            protected_path,
            out_dir,
            out_ext,
            filter_name=filter_name,
            input_filter=(
                f'{FILTER_CSV}:{ord(delimiter)},34,76,1,'
                ',1033,false,false'
            ),
        )
        _assert_spreadsheet_has_no_formulas(produced_path)
        return produced_path
    except BaseException:
        for candidate in {produced_path, expected_path}:
            if not candidate:
                continue
            try:
                os.remove(candidate)
            except OSError:
                pass
        raise
    finally:
        try:
            os.remove(protected_path)
        except OSError:
            pass


def _neutralize_dataframe_for_spreadsheet(dataframe):
    """Cria a barreira de células sem alterar o DataFrame selecionado."""
    safe_dataframe = dataframe.copy(deep=True)
    return safe_dataframe.map(_neutralize_spreadsheet_formula)


def _write_dataframe_to_spreadsheet(
    dataframe,
    writer,
    *,
    sheet_name: str,
) -> None:
    """Único sink de DataFrames extraídos para o workbook PDF -> XLSX."""
    safe_dataframe = _neutralize_dataframe_for_spreadsheet(dataframe)
    safe_headers = [
        _neutralize_spreadsheet_formula(column)
        for column in dataframe.columns
    ]
    safe_dataframe.to_excel(
        writer,
        index=False,
        header=safe_headers,
        sheet_name=sheet_name,
    )


def _excel_safe_str(s: Any) -> str:
    s = "" if s is None else str(s).replace("\r"," ").replace("\n"," ").strip()
    return "'" + s if s.startswith(EXCEL_DANGEROUS_PREFIXES) else s

_num_re = re.compile(r"^[\sR\$\-+()]?[\d\.\,\s]+%?$")
def _maybe_number(s: Any):
    if s is None: return None
    txt = str(s).strip()
    if not txt or not _num_re.match(txt.replace("R$"," ").strip()): return None
    t = txt.replace("R$"," ").replace(" ","").strip()
    neg = False
    if t.startswith("(") and t.endswith(")"):
        neg = True; t = t[1:-1]
    is_percent = t.endswith("%")
    if is_percent: t = t[:-1]
    if "," in t and "." in t:
        t = t.replace(".","").replace(",",".")
    elif "," in t:
        t = t.replace(".","").replace(",",".")
    else:
        t = t.replace(",", "")
    try:
        val = Decimal(t)
        val = -val if neg else val
        if is_percent: val = val / Decimal(100)
        return val
    except InvalidOperation:
        return None


# ---------------- PDF -> XLSX candidate quality ----------------
# A contagem de linhas tem recompensa limitada. Assim, repetir a mesma tabela
# não torna um candidato artificialmente melhor.
PDF_XLSX_ROW_REWARD_CAP = 100
PDF_XLSX_MIN_CANDIDATE_SCORE = 18.0
PDF_XLSX_MASS_DUPLICATION_RATIO = 0.60
PDF_XLSX_MAX_UNIQUE_ROWS_PER_CANDIDATE = 100_000
PDF_XLSX_OVERLAP_IOU = 0.25
PDF_XLSX_OVERLAP_SMALLER_COVERAGE = 0.55

# Uma repetição isolada pode ser um registro legítimo. Blocos só são
# considerados artefatos do extrator quando possuem ao menos três linhas e
# aparecem de forma contígua duas vezes. O limite de 128 inícios candidatos
# mantém a busca determinística e limitada mesmo em tabelas hostis.
PDF_XLSX_REPEAT_BLOCK_MIN_ROWS = 3
PDF_XLSX_REPEAT_BLOCK_MIN_COPIES = 2
PDF_XLSX_REPEAT_BLOCK_MAX_STARTS = 128
PDF_XLSX_HEADER_SAMPLE_ROWS = 5
PDF_XLSX_HEADER_MIN_TYPED_VALUES = 2

PDF_XLSX_HEADER_LABELS = frozenset({
    "apol",
    "apolice",
    "cia",
    "codigo",
    "conta",
    "cpf",
    "data",
    "descricao",
    "documento",
    "fatura",
    "matricula",
    "percentual",
    "procedimento",
    "quantidade",
    "ramo",
    "servico",
    "suc",
    "valor",
})

# A limpeza por frequência é o último recurso: exige duplicação massiva,
# ao menos três fingerprints repetidos três vezes e explosão de 2,5x. Isso
# evita apagar uma segunda ocorrência isolada sem evidência estrutural.
PDF_XLSX_FREQUENCY_MIN_OCCURRENCES = 3
PDF_XLSX_FREQUENCY_MIN_FINGERPRINTS = 3
PDF_XLSX_FREQUENCY_EXPLOSION_FACTOR = 2.5

# A normalização especializada pode aumentar linhas quando realmente separa
# informação. Ela é rejeitada quando o crescimento de linhas é >= 2,5x e o
# crescimento de linhas distintas não acompanha ao menos 60% dessa razão.
PDF_XLSX_POST_NORMALIZATION_ROW_GROWTH = 2.5
PDF_XLSX_POST_NORMALIZATION_UNIQUE_SHARE = 0.60
PDF_XLSX_POST_NORMALIZATION_DUPLICATE_INCREASE = 0.20
PDF_XLSX_POST_NORMALIZATION_DUPLICATE_RATIO = 0.40
PDF_XLSX_POST_NORMALIZATION_MIN_ROW_RETENTION = 0.20
PDF_XLSX_POST_NORMALIZATION_MIN_UNIQUE_RETENTION = 0.35
PDF_XLSX_POST_NORMALIZATION_NEAR_EMPTY_RATIO = 0.50
PDF_XLSX_POST_NORMALIZATION_NEAR_EMPTY_INCREASE = 0.30
PDF_XLSX_POST_NORMALIZATION_COLUMN_GROWTH = 2.0
PDF_XLSX_POST_NORMALIZATION_COLUMN_ALLOWANCE = 12
PDF_XLSX_POST_NORMALIZATION_MIN_DENSITY = 0.20
PDF_XLSX_POST_NORMALIZATION_MIN_DENSITY_RETENTION = 0.35

PDF_XLSX_CAMELOT_WORKER_TIMEOUT_SEC = 60
PDF_XLSX_CAMELOT_WORKER_CLEANUP_MARGIN_SEC = 2.0
PDF_XLSX_CAMELOT_WORKER_MIN_TIMEOUT_SEC = 0.1
PDF_XLSX_CAMELOT_WORKER_MIN_MEM_MB = 256
PDF_XLSX_CAMELOT_WORKER_DEFAULT_MEM_MB = 1024
PDF_XLSX_CAMELOT_WORKER_MAX_MEM_MB = 4096
PDF_XLSX_CAMELOT_WORKER_FILE_MB = 128
PDF_XLSX_CAMELOT_WORKER_MAX_PROCESSES = 16

PDF_XLSX_SCORE_USEFUL_ROWS = 20.0
PDF_XLSX_SCORE_COLUMN_CONSISTENCY = 18.0
PDF_XLSX_SCORE_DENSITY = 15.0
PDF_XLSX_SCORE_PLAUSIBLE_HEADER = 10.0
PDF_XLSX_SCORE_NUMERIC_STABILITY = 8.0
PDF_XLSX_SCORE_COHERENT_FILL = 10.0

PDF_XLSX_PENALTY_DUPLICATION = 85.0
PDF_XLSX_PENALTY_HEADER_REPEAT = 2.0
PDF_XLSX_PENALTY_EMPTY_COLUMNS = 15.0
PDF_XLSX_PENALTY_FRAGMENTATION = 15.0
PDF_XLSX_PENALTY_NEAR_EMPTY_ROWS = 20.0
PDF_XLSX_PENALTY_BOILERPLATE = 15.0
PDF_XLSX_PENALTY_EXPLOSION_MAX = 25.0

PDF_XLSX_EXTRACTOR_PRIORITY = {
    "camelot-lattice-smart": 60,
    "camelot-lattice-region": 55,
    "camelot-lattice-global": 50,
    "camelot-stream-smart": 40,
    "camelot-stream-global": 35,
    "pdfplumber": 25,
    "text": 10,
}

_PDF_XLSX_BOILERPLATE_RE = re.compile(
    r"(?:^\s*p[aá]gina\s+\d+\s*$|"
    r"\b(?:observa[cç][aã]o|obs\.?|anexo|rodap[eé]|confidencial)\b)",
    re.IGNORECASE,
)


@dataclass
class TableCandidate:
    """Candidato interno de tabela com procedência e métricas não sensíveis."""

    dataframe: Any
    page_number: int
    extractor: str
    bbox: Optional[Tuple[float, float, float, float]] = None
    region_id: str = ""
    group_id: str = ""
    scope: str = "region"
    raw_rows: int = 0
    raw_columns: int = 0
    useful_rows: int = 0
    useful_columns: int = 0
    unique_rows: int = 0
    density: float = 0.0
    duplicate_rows: int = 0
    duplicate_ratio: float = 0.0
    fully_empty_columns: int = 0
    repeated_header_rows: int = 0
    column_consistency: float = 0.0
    plausible_header: float = 0.0
    numeric_stability: float = 0.0
    near_empty_ratio: float = 0.0
    fragmentation_ratio: float = 0.0
    boilerplate_ratio: float = 0.0
    score: float = 0.0
    duplicate_rows_removed: int = 0
    repeated_block_rows_removed: int = 0
    repeated_header_rows_removed: int = 0
    frequency_rows_removed: int = 0
    structural_report: Optional[Dict[str, Any]] = None
    geometry_report: Optional[CandidateGeometryReport] = None


@dataclass(frozen=True)
class NormalizationAssessment:
    accepted: bool
    reasons: Tuple[str, ...]
    before: Dict[str, Any]
    after: Dict[str, Any]


@dataclass
class PreparedWorkbookTable:
    dataframe: Any
    meta: Dict[str, Dict[str, Any]]
    sheet_name: str
    page_number: int
    group_id: str
    extractor: str
    normalization_accepted: bool = False


def _comparison_cell(value: Any) -> str:
    """
    Normalização somente para comparação/fingerprint.

    Não altera o DataFrame exportado e evita conversão agressiva de códigos:
    inteiros textuais e valores com ponto ambíguo permanecem texto.
    """
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, Decimal) and value.is_nan():
        return ""
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        try:
            number = Decimal(str(value))
            if number.is_finite():
                normalized_number = format(number.normalize(), "f")
                if normalized_number == "-0":
                    normalized_number = "0"
                return f"number:{normalized_number}"
        except (InvalidOperation, ValueError):
            pass

    text = unicodedata.normalize("NFKC", str(value))
    text = re.sub(r"\s+", " ", text.replace("\r", " ").replace("\n", " ")).strip()
    if not text or text.lower() in {"nan", "none", "<na>"}:
        return ""

    explicit_numeric = (
        "R$" in text
        or text.endswith("%")
        or bool(re.fullmatch(r"[+-]?\d{1,3}(?:\.\d{3})*,\d+", text))
        or bool(re.fullmatch(r"[+-]?\d+,\d+", text))
    )
    if explicit_numeric:
        parsed = _maybe_number(text)
        if parsed is not None:
            normalized = format(parsed.normalize(), "f")
            if normalized == "-0":
                normalized = "0"
            kind = "percent" if text.endswith("%") else (
                "currency" if "R$" in text else "number"
            )
            return f"{kind}:{normalized}"
    return text


def _candidate_normalized_rows(df: Any) -> Tuple[List[List[str]], List[int]]:
    try:
        raw_rows = list(df.itertuples(index=False, name=None))
        raw_columns = int(df.shape[1])
    except Exception:
        return [], []

    normalized = [
        [_comparison_cell(value) for value in row]
        for row in raw_rows
    ]
    useful_columns = [
        index
        for index in range(raw_columns)
        if any(index < len(row) and row[index] for row in normalized)
    ]
    return normalized, useful_columns


def _candidate_row_fingerprints(
    df: Any,
) -> Tuple[List[Tuple[str, ...]], List[int], List[List[str]]]:
    normalized, useful_columns = _candidate_normalized_rows(df)
    fingerprints = [
        tuple(row[index] if index < len(row) else "" for index in useful_columns)
        for row in normalized
    ]
    return fingerprints, useful_columns, normalized


def _numeric_stability(
    normalized_rows: Sequence[Sequence[str]],
    useful_columns: Sequence[int],
) -> float:
    if not useful_columns:
        return 0.0
    per_column: List[float] = []
    for index in useful_columns:
        values = [
            row[index]
            for row in normalized_rows
            if index < len(row) and row[index]
        ]
        if not values:
            continue
        numeric = sum(
            value.startswith(("number:", "currency:", "percent:"))
            for value in values
        )
        textual = len(values) - numeric
        per_column.append(max(numeric, textual) / len(values))
    return sum(per_column) / len(per_column) if per_column else 0.0


def _table_candidate_metrics(df: Any) -> Dict[str, Any]:
    try:
        raw_rows = int(df.shape[0])
        raw_columns = int(df.shape[1])
    except Exception:
        raw_rows = raw_columns = 0

    fingerprints, useful_column_indexes, normalized = (
        _candidate_row_fingerprints(df)
    )
    useful_fingerprints = [
        fingerprint for fingerprint in fingerprints if any(fingerprint)
    ]
    useful_rows = len(useful_fingerprints)
    useful_columns = len(useful_column_indexes)
    counts = Counter(useful_fingerprints)
    duplicate_rows = sum(count - 1 for count in counts.values() if count > 1)
    unique_rows = len(counts)
    duplicate_ratio = duplicate_rows / useful_rows if useful_rows else 0.0

    filled_counts = [sum(bool(cell) for cell in row) for row in useful_fingerprints]
    filled_total = sum(filled_counts)
    density = (
        filled_total / (useful_rows * useful_columns)
        if useful_rows and useful_columns
        else 0.0
    )
    mean_fill = (
        sum(filled_counts) / len(filled_counts)
        if filled_counts
        else 0.0
    )
    column_consistency = 0.0
    if filled_counts and useful_columns:
        mean_deviation = sum(
            abs(value - mean_fill) for value in filled_counts
        ) / len(filled_counts)
        column_consistency = max(0.0, 1.0 - (mean_deviation / useful_columns))

    header_fingerprint: Optional[Tuple[str, ...]] = None
    if useful_fingerprints:
        header_fingerprint = max(
            useful_fingerprints[: min(5, len(useful_fingerprints))],
            key=lambda row: (sum(bool(cell) for cell in row), -len("".join(row))),
        )
    repeated_header_rows = (
        max(0, counts.get(header_fingerprint, 0) - 1)
        if header_fingerprint
        else 0
    )
    plausible_header = 0.0
    if header_fingerprint:
        nonempty_header = [cell for cell in header_fingerprint if cell]
        alpha_cells = sum(
            any(character.isalpha() for character in cell)
            and not cell.startswith(("number:", "currency:", "percent:"))
            for cell in nonempty_header
        )
        if nonempty_header:
            plausible_header = alpha_cells / len(nonempty_header)
            if len(set(nonempty_header)) != len(nonempty_header):
                plausible_header *= 0.75

    near_empty_limit = max(1, math.ceil(max(1, useful_columns) * 0.25))
    near_empty_rows = sum(value <= near_empty_limit for value in filled_counts)
    near_empty_ratio = near_empty_rows / useful_rows if useful_rows else 0.0

    fragmented_cells = 0
    nonempty_cells = 0
    boilerplate_rows = 0
    try:
        raw_values = list(df.itertuples(index=False, name=None))
    except Exception:
        raw_values = []
    for row in raw_values:
        row_text_parts = []
        for value in row:
            if value is None:
                continue
            text = str(value)
            if not text.strip():
                continue
            nonempty_cells += 1
            row_text_parts.append(text.strip())
            if "\n" in text or "\r" in text or len(text) > 240:
                fragmented_cells += 1
        row_text = " ".join(row_text_parts)
        if row_text and _PDF_XLSX_BOILERPLATE_RE.search(row_text):
            boilerplate_rows += 1

    fragmentation_ratio = (
        fragmented_cells / nonempty_cells if nonempty_cells else 0.0
    )
    boilerplate_ratio = (
        boilerplate_rows / useful_rows if useful_rows else 0.0
    )
    numeric_stability = _numeric_stability(normalized, useful_column_indexes)

    return {
        "raw_rows": raw_rows,
        "raw_columns": raw_columns,
        "useful_rows": useful_rows,
        "useful_columns": useful_columns,
        "unique_rows": unique_rows,
        "density": density,
        "duplicate_rows": duplicate_rows,
        "duplicate_ratio": duplicate_ratio,
        "fully_empty_columns": max(0, raw_columns - useful_columns),
        "repeated_header_rows": repeated_header_rows,
        "column_consistency": column_consistency,
        "plausible_header": plausible_header,
        "numeric_stability": numeric_stability,
        "near_empty_ratio": near_empty_ratio,
        "fragmentation_ratio": fragmentation_ratio,
        "boilerplate_ratio": boilerplate_ratio,
    }


def _score_table_candidate(candidate: TableCandidate) -> float:
    unique_row_reward = (
        min(candidate.unique_rows, PDF_XLSX_ROW_REWARD_CAP)
        / PDF_XLSX_ROW_REWARD_CAP
    ) * PDF_XLSX_SCORE_USEFUL_ROWS
    empty_column_ratio = (
        candidate.fully_empty_columns / candidate.raw_columns
        if candidate.raw_columns
        else 1.0
    )
    explosion_factor = (
        candidate.useful_rows / max(1, candidate.unique_rows)
        if candidate.useful_rows
        else 1.0
    )
    explosion_penalty = min(
        PDF_XLSX_PENALTY_EXPLOSION_MAX,
        max(0.0, explosion_factor - 1.5) * 8.0,
    )

    score = (
        unique_row_reward
        + candidate.column_consistency * PDF_XLSX_SCORE_COLUMN_CONSISTENCY
        + candidate.density * PDF_XLSX_SCORE_DENSITY
        + candidate.plausible_header * PDF_XLSX_SCORE_PLAUSIBLE_HEADER
        + candidate.numeric_stability * PDF_XLSX_SCORE_NUMERIC_STABILITY
        + (1.0 - candidate.near_empty_ratio) * PDF_XLSX_SCORE_COHERENT_FILL
        - candidate.duplicate_ratio * PDF_XLSX_PENALTY_DUPLICATION
        - min(candidate.repeated_header_rows, 10)
        * PDF_XLSX_PENALTY_HEADER_REPEAT
        - empty_column_ratio * PDF_XLSX_PENALTY_EMPTY_COLUMNS
        - candidate.fragmentation_ratio * PDF_XLSX_PENALTY_FRAGMENTATION
        - candidate.near_empty_ratio * PDF_XLSX_PENALTY_NEAR_EMPTY_ROWS
        - candidate.boilerplate_ratio * PDF_XLSX_PENALTY_BOILERPLATE
        - explosion_penalty
    )
    return round(score, 6)


def _make_table_candidate(
    dataframe: Any,
    *,
    page_number: int,
    extractor: str,
    bbox: Optional[Tuple[float, float, float, float]] = None,
    region_id: str = "",
    scope: str = "region",
    structural_report: Optional[Dict[str, Any]] = None,
) -> TableCandidate:
    candidate = TableCandidate(
        dataframe=dataframe,
        page_number=max(1, int(page_number or 1)),
        extractor=extractor,
        bbox=_normalize_candidate_bbox(bbox),
        region_id=region_id,
        scope=scope,
        structural_report=(
            dict(structural_report)
            if structural_report is not None
            else None
        ),
    )
    for key, value in _table_candidate_metrics(dataframe).items():
        setattr(candidate, key, value)
    candidate.score = _score_table_candidate(candidate)
    return candidate


def _normalize_candidate_bbox(
    bbox: Optional[Sequence[float]],
) -> Optional[Tuple[float, float, float, float]]:
    if bbox is None or len(bbox) != 4:
        return None
    try:
        x0, y0, x1, y1 = (float(value) for value in bbox)
    except (TypeError, ValueError):
        return None
    left, right = sorted((x0, x1))
    bottom, top = sorted((y0, y1))
    if right <= left or top <= bottom:
        return None
    return left, bottom, right, top


def _candidate_bbox_from_area(
    area: Optional[str],
) -> Optional[Tuple[float, float, float, float]]:
    if not area:
        return None
    try:
        return _normalize_candidate_bbox(
            [part.strip() for part in area.split(",")]
        )
    except Exception:
        return None


def _map_pdf_table_geometry_shadow(
    src_pdf: str,
    expected_pages: Sequence[int],
) -> Dict[int, PageTableGeometry]:
    """Produz somente metadados; qualquer falha geométrica mantém o fluxo atual."""
    try:
        mapped = map_pdf_table_geometry(
            src_pdf,
            expected_pages,
            check_deadline=check_converter_deadline,
        )
    except ConverterTimeoutError:
        raise
    except Exception as exc:
        logger.warning(
            "PDF->XLSX geometry stage=failed error=%s",
            type(exc).__name__,
        )
        return {}
    for page_number in sorted(mapped):
        page = mapped[page_number]
        logger.info(
            "PDF->XLSX geometry stage=mapped page=%d regions=%d "
            "column_bands=%d row_bands=%d section_count=%d "
            "data_row_count=%d total_count=%d note_count=%d "
            "geometry_confidence=%.4f limited=%s",
            page_number,
            len(page.regions),
            sum(len(region.columns) for region in page.regions),
            sum(len(region.rows) for region in page.regions),
            sum(region.role_count("section") for region in page.regions),
            sum(region.role_count("data") for region in page.regions),
            sum(region.role_count("total") for region in page.regions),
            page.note_count
            + sum(region.role_count("note") for region in page.regions),
            page.confidence,
            str(page.limited).lower(),
        )
    return mapped


def _associate_geometry_reports(
    candidates: Sequence[TableCandidate],
    geometry_by_page: Dict[int, PageTableGeometry],
) -> None:
    """Associa o mapa sem tocar no DataFrame, score ou procedência existente."""
    for candidate in candidates:
        check_converter_deadline("pdf-xlsx-geometry-association")
        page = geometry_by_page.get(candidate.page_number)
        report = geometry_report_for_candidate(page, candidate.bbox)
        candidate.geometry_report = report
        logger.info(
            "PDF->XLSX geometry stage=association page=%d mapped=%s "
            "ambiguous=%s column_bands=%d row_bands=%d "
            "geometry_confidence=%.4f",
            candidate.page_number,
            str(report.mapped).lower(),
            str(report.ambiguous).lower(),
            report.column_band_count,
            report.row_band_count,
            report.geometry_confidence,
        )


def _camelot_worker_budget(stage: str) -> Tuple[float, bool]:
    configured = _positive_env_int(
        ("PDF_TO_XLSX_CAMELOT_TIMEOUT_SEC",),
        PDF_XLSX_CAMELOT_WORKER_TIMEOUT_SEC,
    )
    runtime = _ACTIVE_CONVERTER_RUNTIME.get()
    if runtime is None:
        return float(configured), False
    remaining = runtime.remaining(stage)
    usable = remaining - PDF_XLSX_CAMELOT_WORKER_CLEANUP_MARGIN_SEC
    if usable < PDF_XLSX_CAMELOT_WORKER_MIN_TIMEOUT_SEC:
        raise _ConverterDeadlineReservedError(
            "Orçamento insuficiente para iniciar o worker Camelot."
        )
    operation_timeout = float(configured)
    return min(operation_timeout, usable), usable <= operation_timeout


def _camelot_worker_memory_mb() -> int:
    configured = _positive_env_int(
        ("PDF_TO_XLSX_CAMELOT_MEM_MB",),
        PDF_XLSX_CAMELOT_WORKER_DEFAULT_MEM_MB,
    )
    return min(
        PDF_XLSX_CAMELOT_WORKER_MAX_MEM_MB,
        max(PDF_XLSX_CAMELOT_WORKER_MIN_MEM_MB, configured),
    )


def _camelot_worker_env(attempt_dir: str) -> Dict[str, str]:
    allowed_names = (
        "PATH",
        "PATHEXT",
        "SYSTEMROOT",
        "WINDIR",
        "COMSPEC",
        "LANG",
        "LC_ALL",
        "LD_LIBRARY_PATH",
        "DYLD_LIBRARY_PATH",
        "GHOSTSCRIPT_PATH",
        "GS_PROG",
        "GS_BIN",
        "GHOSTSCRIPT_BIN",
        "PYTHONIOENCODING",
        "PYTHONUTF8",
    )
    worker_env = {
        name: os.environ[name]
        for name in allowed_names
        if os.environ.get(name)
    }
    worker_env.update({
        "TEMP": attempt_dir,
        "TMP": attempt_dir,
        "TMPDIR": attempt_dir,
        "MPLCONFIGDIR": attempt_dir,
        "XDG_CACHE_HOME": attempt_dir,
        "XDG_CONFIG_HOME": attempt_dir,
        "PYTHONIOENCODING": "utf-8",
        "PYTHONUTF8": "1",
    })
    return worker_env


def _path_is_inside(parent: str, child: str) -> bool:
    parent_real = os.path.realpath(os.path.abspath(parent))
    child_real = os.path.realpath(os.path.abspath(child))
    try:
        return (
            child_real != parent_real
            and os.path.commonpath((parent_real, child_real))
            == parent_real
        )
    except ValueError:
        return False


def _write_camelot_request_atomic(
    path: str,
    payload: Dict[str, Any],
) -> None:
    encoded = json.dumps(
        payload,
        ensure_ascii=False,
        allow_nan=False,
        separators=(",", ":"),
    ).encode("utf-8")
    if not encoded or len(encoded) > _camelot_worker.MAX_REQUEST_BYTES:
        raise ConverterToolExecutionError(
            "Solicitação interna do Camelot excede o limite."
        )
    fd, temporary = tempfile.mkstemp(
        prefix=".camelot-request-",
        suffix=".tmp",
        dir=os.path.dirname(path),
    )
    try:
        try:
            os.chmod(temporary, 0o600)
        except OSError:
            pass
        with os.fdopen(fd, "wb") as stream:
            stream.write(encoded)
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temporary, path)
    finally:
        try:
            if os.path.exists(temporary):
                os.remove(temporary)
        except OSError:
            pass


def _validate_camelot_worker_response(
    result_path: str,
    workdir: str,
    request_payload: Dict[str, Any],
) -> List[TableCandidate]:
    if (
        not _path_is_inside(workdir, result_path)
        or os.path.islink(result_path)
        or not os.path.isfile(result_path)
    ):
        raise ConverterToolExecutionError(
            "Resultado interno do Camelot fora do workdir."
        )
    try:
        result_size = os.path.getsize(result_path)
    except OSError as exc:
        raise ConverterToolExecutionError(
            "Resultado interno do Camelot ilegível."
        ) from exc
    if (
        result_size <= 0
        or result_size > _camelot_worker.MAX_RESULT_BYTES
    ):
        raise ConverterToolExecutionError(
            "Resultado interno do Camelot excede o limite."
        )
    try:
        with open(
            result_path,
            "r",
            encoding="utf-8",
            errors="strict",
        ) as stream:
            response = json.load(stream)
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise ConverterToolExecutionError(
            "Resultado JSON interno do Camelot inválido."
        ) from exc
    if not isinstance(response, dict):
        raise ConverterToolExecutionError(
            "Resposta interna do Camelot inválida."
        )
    if set(response) != set(_camelot_worker._RESPONSE_KEYS):
        raise ConverterToolExecutionError(
            "Resposta interna do Camelot possui campos inválidos."
        )
    for key in (
        "protocol",
        "operation",
        "request_id",
        "flavor",
        "pages",
        "page_hint",
        "extractor",
        "region_prefix",
        "region_index_width",
        "table_area",
    ):
        if response.get(key) != request_payload.get(key):
            raise ConverterToolExecutionError(
                "Resposta interna do Camelot não corresponde à solicitação."
            )

    tables = response.get("tables")
    if (
        not isinstance(tables, list)
        or len(tables) > _camelot_worker.MAX_TABLES
    ):
        raise ConverterToolExecutionError(
            "Quantidade de tabelas do Camelot inválida."
        )
    allowed_pages = _camelot_worker.parse_pages_spec(
        request_payload["pages"]
    )
    total_rows = 0
    total_cells = 0
    candidates: List[TableCandidate] = []
    import pandas as pd

    for table_index, table in enumerate(tables, start=1):
        if (
            not isinstance(table, dict)
            or set(table) != set(_camelot_worker._TABLE_KEYS)
        ):
            raise ConverterToolExecutionError(
                "Tabela interna do Camelot inválida."
            )
        page = table.get("page")
        row_count = table.get("row_count")
        column_count = table.get("column_count")
        if (
            isinstance(page, bool)
            or not isinstance(page, int)
            or page not in allowed_pages
            or isinstance(row_count, bool)
            or not isinstance(row_count, int)
            or row_count <= 0
            or isinstance(column_count, bool)
            or not isinstance(column_count, int)
            or column_count <= 0
            or column_count > _camelot_worker.MAX_COLUMNS
        ):
            raise ConverterToolExecutionError(
                "Dimensões internas do Camelot inválidas."
            )
        total_rows += row_count
        total_cells += row_count * column_count
        if (
            total_rows > _camelot_worker.MAX_TOTAL_ROWS
            or total_cells > _camelot_worker.MAX_TOTAL_CELLS
        ):
            raise ConverterToolExecutionError(
                "Resultado interno do Camelot excessivo."
            )
        rows = table.get("rows")
        if not isinstance(rows, list) or len(rows) != row_count:
            raise ConverterToolExecutionError(
                "Linhas internas do Camelot inválidas."
            )
        for row in rows:
            if not isinstance(row, list) or len(row) != column_count:
                raise ConverterToolExecutionError(
                    "Linha interna do Camelot irregular."
                )
            for value in row:
                if value is not None and not isinstance(
                    value,
                    (str, int, float, bool),
                ):
                    raise ConverterToolExecutionError(
                        "Tipo de célula interno inválido."
                    )
                if (
                    isinstance(value, float)
                    and not math.isfinite(value)
                ):
                    raise ConverterToolExecutionError(
                        "Número interno inválido."
                    )
                if (
                    isinstance(value, str)
                    and len(value) > _camelot_worker.MAX_CELL_CHARS
                ):
                    raise ConverterToolExecutionError(
                        "Célula interna excede o limite."
                    )

        bbox = table.get("bbox")
        if bbox is not None:
            if (
                not isinstance(bbox, list)
                or len(bbox) != 4
                or any(
                    isinstance(value, bool)
                    or not isinstance(value, (int, float))
                    or not math.isfinite(float(value))
                    or abs(float(value)) > 1_000_000
                    for value in bbox
                )
            ):
                raise ConverterToolExecutionError(
                    "BBox interna do Camelot inválida."
                )
        normalized_bbox = _normalize_candidate_bbox(bbox)
        if bbox is not None and normalized_bbox is None:
            raise ConverterToolExecutionError(
                "BBox interna do Camelot é degenerada."
            )
        requested_bbox = _candidate_bbox_from_area(
            request_payload.get("table_area")
        )
        if (
            normalized_bbox is not None
            and requested_bbox is not None
            and not _bbox_materially_overlaps(
                normalized_bbox,
                requested_bbox,
            )
        ):
            raise ConverterToolExecutionError(
                "BBox interna do Camelot não corresponde à região."
            )
        if normalized_bbox is None:
            normalized_bbox = _candidate_bbox_from_area(
                request_payload.get("table_area")
            )

        report = table.get("report")
        if (
            not isinstance(report, dict)
            or set(report) - set(_camelot_worker._REPORT_KEYS)
            or any(
                isinstance(value, bool)
                or not isinstance(value, (int, float))
                or not math.isfinite(float(value))
                or abs(float(value)) > 1_000_000
                for value in report.values()
            )
        ):
            raise ConverterToolExecutionError(
                "Relatório estrutural do Camelot inválido."
            )
        if "page" in report and float(report["page"]) != float(page):
            raise ConverterToolExecutionError(
                "Relatório estrutural do Camelot possui página inválida."
            )

        dataframe = pd.DataFrame(rows)
        candidate = _make_table_candidate(
            dataframe,
            page_number=page,
            extractor=request_payload["extractor"],
            bbox=normalized_bbox,
            region_id=(
                f"{request_payload['region_prefix']}-"
                f"{table_index:0{request_payload['region_index_width']}d}"
            ),
            scope="region" if normalized_bbox is not None else "page",
            structural_report=dict(report),
        )
        candidates.append(candidate)
    return candidates


def _run_camelot_worker(
    src_pdf: str,
    *,
    flavor: str,
    pages: str,
    page_hint: int,
    extractor: str,
    region_prefix: str,
    region_index_width: int,
    table_area: Optional[str] = None,
    columns: Optional[str] = None,
    line_scale: Optional[int] = None,
    strip_text: str = "\n",
    process_background: Optional[bool] = None,
    copy_text: Optional[List[str]] = None,
    shift_text: Optional[List[str]] = None,
    dpi: int = 200,
) -> List[TableCandidate]:
    source_abs = os.path.abspath(src_pdf)
    workdir = os.path.dirname(source_abs)
    if (
        not os.path.isfile(source_abs)
        or os.path.islink(source_abs)
        or os.path.basename(source_abs) in {"", ".", ".."}
        or not _path_is_inside(workdir, source_abs)
    ):
        raise ConverterToolExecutionError(
            "Entrada interna do Camelot inválida."
        )

    request_id = secrets.token_hex(16)
    options: Dict[str, Any] = {
        "strip_text": strip_text,
        "dpi": int(dpi),
    }
    if flavor == "lattice":
        options.update({
            "line_scale": int(line_scale or 80),
            "process_background": bool(process_background),
            "copy_text": list(copy_text or ["h", "v"]),
            "shift_text": list(shift_text or ["l", "t"]),
        })
    elif columns is not None:
        options["columns"] = columns
    request_payload: Dict[str, Any] = {
        "protocol": _camelot_worker.PROTOCOL_VERSION,
        "operation": "extract",
        "request_id": request_id,
        "input_file": os.path.basename(source_abs),
        "flavor": flavor,
        "pages": pages,
        "page_hint": int(page_hint),
        "extractor": extractor,
        "region_prefix": region_prefix,
        "region_index_width": int(region_index_width),
        "table_area": table_area,
        "options": options,
    }
    try:
        _camelot_worker.validate_request(
            request_payload,
            workdir,
        )
    except _camelot_worker.WorkerRequestError as exc:
        raise ConverterToolExecutionError(
            "Solicitação interna do Camelot inválida."
        ) from exc

    timeout, deadline_limited = _camelot_worker_budget(
        f"camelot-{flavor}-{page_hint}"
    )
    attempt_name = f".camelot-worker-{request_id}"
    attempt_dir = os.path.join(workdir, attempt_name)
    if not _path_is_inside(workdir, attempt_dir):
        raise ConverterToolExecutionError(
            "Diretório interno do Camelot inválido."
        )
    os.makedirs(attempt_dir, mode=0o700, exist_ok=False)
    request_path = os.path.join(attempt_dir, "request.json")
    result_path = os.path.join(attempt_dir, "result.json")
    request_rel = os.path.join(attempt_name, "request.json")
    result_rel = os.path.join(attempt_name, "result.json")
    started_at = time.monotonic()
    try:
        _write_camelot_request_atomic(
            request_path,
            request_payload,
        )
        worker_script = os.path.realpath(
            os.path.abspath(_camelot_worker.__file__)
        )
        command = [
            sys.executable,
            worker_script,
            "--request",
            request_rel,
            "--result",
            result_rel,
        ]
        logger.info(
            "PDF->XLSX camelot_worker stage=start operation=extract "
            "flavor=%s page=%d region=%s timeout=%.3f",
            flavor,
            page_hint,
            region_prefix,
            timeout,
        )
        try:
            result = run_in_sandbox(
                command,
                cwd=workdir,
                timeout=timeout,
                cpu_seconds=max(1, math.ceil(timeout)),
                mem_mb=_camelot_worker_memory_mb(),
                env=_camelot_worker_env(attempt_dir),
                file_mb=PDF_XLSX_CAMELOT_WORKER_FILE_MB,
                max_processes=PDF_XLSX_CAMELOT_WORKER_MAX_PROCESSES,
                output_limit_chars=2_048,
            )
        except subprocess.TimeoutExpired:
            logger.warning(
                "PDF->XLSX camelot_worker stage=timeout operation=extract "
                "flavor=%s page=%d region=%s timeout=%.3f",
                flavor,
                page_hint,
                region_prefix,
                timeout,
            )
            if deadline_limited:
                raise _ConverterDeadlineReservedError(
                    "Worker Camelot consumiu o orçamento restante."
                ) from None
            raise ConverterTimeoutError(
                "Worker Camelot excedeu o tempo permitido."
            ) from None
        if result.returncode != _camelot_worker.EXIT_OK:
            logger.warning(
                "PDF->XLSX camelot_worker stage=failed operation=extract "
                "flavor=%s page=%d region=%s returncode=%d",
                flavor,
                page_hint,
                region_prefix,
                int(result.returncode),
            )
            raise ConverterToolExecutionError(
                "Worker Camelot terminou com erro."
            )
        candidates = _validate_camelot_worker_response(
            result_path,
            workdir,
            request_payload,
        )
        logger.info(
            "PDF->XLSX camelot_worker stage=complete operation=extract "
            "flavor=%s page=%d region=%s elapsed=%.3f candidates=%d "
            "returncode=%d",
            flavor,
            page_hint,
            region_prefix,
            time.monotonic() - started_at,
            len(candidates),
            int(result.returncode),
        )
        return candidates
    finally:
        if _path_is_inside(workdir, attempt_dir):
            shutil.rmtree(attempt_dir, ignore_errors=True)


def _bbox_materially_overlaps(
    first: Optional[Tuple[float, float, float, float]],
    second: Optional[Tuple[float, float, float, float]],
) -> bool:
    if first is None or second is None:
        return False
    left = max(first[0], second[0])
    bottom = max(first[1], second[1])
    right = min(first[2], second[2])
    top = min(first[3], second[3])
    if right <= left or top <= bottom:
        return False
    intersection = (right - left) * (top - bottom)
    first_area = (first[2] - first[0]) * (first[3] - first[1])
    second_area = (second[2] - second[0]) * (second[3] - second[1])
    union = first_area + second_area - intersection
    iou = intersection / union if union else 0.0
    smaller_coverage = intersection / min(first_area, second_area)
    return (
        iou >= PDF_XLSX_OVERLAP_IOU
        or smaller_coverage >= PDF_XLSX_OVERLAP_SMALLER_COVERAGE
    )


def _candidate_reading_order(candidate: TableCandidate) -> Tuple[Any, ...]:
    if candidate.bbox is None:
        return candidate.page_number, 1, 0.0, 0.0, candidate.region_id
    left, _bottom, _right, top = candidate.bbox
    return candidate.page_number, 0, -top, left, candidate.region_id


def _group_competing_candidates(
    candidates: Sequence[TableCandidate],
) -> List[List[TableCandidate]]:
    """
    Agrupa por página e pela bbox âncora.

    A bbox do grupo não é expandida; isso evita o encadeamento transitivo no
    qual uma região grande conectaria duas tabelas independentes.
    """
    groups: List[List[TableCandidate]] = []
    anchors: List[Optional[Tuple[float, float, float, float]]] = []
    for candidate in sorted(candidates, key=_candidate_reading_order):
        group_index: Optional[int] = None
        for index, group in enumerate(groups):
            if group[0].page_number != candidate.page_number:
                continue
            anchor = anchors[index]
            if candidate.bbox is None and anchor is None:
                group_index = index
                break
            if _bbox_materially_overlaps(anchor, candidate.bbox):
                group_index = index
                break
        if group_index is None:
            groups.append([candidate])
            anchors.append(candidate.bbox)
        else:
            groups[group_index].append(candidate)

    page_group_counts: Dict[int, int] = {}
    for group in groups:
        page = group[0].page_number
        page_group_counts[page] = page_group_counts.get(page, 0) + 1
        group_id = f"p{page:04d}-g{page_group_counts[page]:03d}"
        for candidate in group:
            candidate.group_id = group_id
    return groups


def _candidate_rank(candidate: TableCandidate) -> Tuple[Any, ...]:
    return (
        candidate.score,
        -candidate.duplicate_ratio,
        candidate.column_consistency,
        candidate.density,
        -candidate.fragmentation_ratio,
        PDF_XLSX_EXTRACTOR_PRIORITY.get(candidate.extractor, 0),
        -candidate.raw_rows,
        candidate.extractor,
        candidate.region_id,
    )


def _select_candidate_winners(
    candidates: Sequence[TableCandidate],
) -> List[TableCandidate]:
    groups = _group_competing_candidates(candidates)
    winners_by_page: Dict[int, Dict[str, List[TableCandidate]]] = {}
    for group in groups:
        winner = max(group, key=_candidate_rank)
        bucket = winners_by_page.setdefault(
            winner.page_number,
            {"region": [], "page": []},
        )
        bucket["region" if winner.bbox is not None else "page"].append(winner)

    selected: List[TableCandidate] = []
    for page_number in sorted(winners_by_page):
        buckets = winners_by_page[page_number]
        region_winners = buckets["region"]
        page_winners = buckets["page"]
        if region_winners:
            if any(
                candidate.score >= PDF_XLSX_MIN_CANDIDATE_SCORE
                for candidate in region_winners
            ):
                selected.extend(region_winners)
            elif page_winners:
                best_region = max(region_winners, key=_candidate_rank)
                best_page = max(page_winners, key=_candidate_rank)
                selected.append(max((best_region, best_page), key=_candidate_rank))
            else:
                selected.extend(region_winners)
        elif page_winners:
            selected.append(max(page_winners, key=_candidate_rank))
    return sorted(selected, key=_candidate_reading_order)


def _recognized_header_fingerprint(
    fingerprints: Sequence[Tuple[str, ...]],
) -> Optional[Tuple[str, ...]]:
    """
    Reconhece somente a primeira linha útil quando há evidência de cabeçalho.

    Texto alfabético isolado não basta: ele poderia ser um registro legítimo.
    Exigimos rótulos tabulares conhecidos ou contraste com valores tipados nas
    linhas seguintes. Em caso ambíguo, preservamos a repetição.
    """
    if not fingerprints:
        return None
    fingerprint = fingerprints[0]
    nonempty = [cell for cell in fingerprint if cell]
    if len(nonempty) < 2:
        return None
    if any(
        cell.startswith(("number:", "currency:", "percent:"))
        for cell in nonempty
    ):
        return None
    alpha_cells = sum(
        any(character.isalpha() for character in cell)
        for cell in nonempty
    )
    if alpha_cells / len(nonempty) < 0.60:
        return None

    normalized_labels = set()
    for cell in nonempty:
        decomposed = unicodedata.normalize("NFKD", cell)
        without_accents = "".join(
            character
            for character in decomposed
            if not unicodedata.combining(character)
        )
        normalized_labels.add(
            re.sub(r"[^a-z0-9]+", "", without_accents.lower())
        )
    label_evidence = len(
        normalized_labels & PDF_XLSX_HEADER_LABELS
    ) >= 2

    typed_evidence = 0
    for column_index, header_cell in enumerate(fingerprint):
        if not header_cell:
            continue
        typed_values = sum(
            row[column_index].startswith(
                ("number:", "currency:", "percent:")
            )
            for row in fingerprints[1:1 + PDF_XLSX_HEADER_SAMPLE_ROWS]
            if column_index < len(row) and row[column_index]
        )
        if typed_values >= PDF_XLSX_HEADER_MIN_TYPED_VALUES:
            typed_evidence += 1

    return fingerprint if label_evidence or typed_evidence else None


def _repeated_block_duplicate_positions(
    fingerprints: Sequence[Tuple[str, ...]],
) -> set[int]:
    """
    Localiza cópias contíguas de blocos, inclusive uma última cópia parcial.

    A busca usa apenas posições que compartilham o primeiro fingerprint do
    bloco e limita os inícios examinados para evitar custo não limitado.
    """
    positions_by_fingerprint: Dict[Tuple[str, ...], List[int]] = {}
    for position, fingerprint in enumerate(fingerprints):
        positions_by_fingerprint.setdefault(fingerprint, []).append(position)

    removals: set[int] = set()
    total = len(fingerprints)
    for start in range(total):
        possible_starts = positions_by_fingerprint.get(
            fingerprints[start],
            [],
        )
        candidates_checked = 0
        best: Optional[Tuple[int, int, int]] = None
        for repeated_start in possible_starts:
            if repeated_start <= start:
                continue
            block_size = repeated_start - start
            if block_size < PDF_XLSX_REPEAT_BLOCK_MIN_ROWS:
                continue
            if repeated_start + block_size > total:
                break
            candidates_checked += 1
            if candidates_checked > PDF_XLSX_REPEAT_BLOCK_MAX_STARTS:
                break
            block = fingerprints[start:repeated_start]
            if (
                fingerprints[
                    repeated_start:repeated_start + block_size
                ]
                != block
            ):
                continue

            copies = 1
            cursor = repeated_start
            while (
                cursor + block_size <= total
                and fingerprints[cursor:cursor + block_size] == block
            ):
                copies += 1
                cursor += block_size
            if copies < PDF_XLSX_REPEAT_BLOCK_MIN_COPIES:
                continue

            partial = 0
            while (
                cursor + partial < total
                and partial < block_size
                and fingerprints[cursor + partial] == block[partial]
            ):
                partial += 1
            if partial < PDF_XLSX_REPEAT_BLOCK_MIN_ROWS:
                partial = 0

            removable = ((copies - 1) * block_size) + partial
            candidate = (removable, -block_size, cursor + partial)
            if best is None or candidate > best:
                best = candidate

        if best is None:
            continue
        removable, negative_block_size, end = best
        block_size = -negative_block_size
        first_duplicate = start + block_size
        removals.update(range(first_duplicate, end))
        if removable <= 0:
            continue
    return removals


def _deduplicate_candidate_rows(candidate: TableCandidate) -> TableCandidate:
    fingerprints, useful_columns, _normalized = _candidate_row_fingerprints(
        candidate.dataframe
    )
    if not fingerprints or not useful_columns:
        return candidate

    useful_entries = [
        (position, fingerprint)
        for position, fingerprint in enumerate(fingerprints)
        if any(fingerprint)
    ]
    useful_fingerprints = [
        fingerprint for _position, fingerprint in useful_entries
    ]
    if not useful_fingerprints:
        return candidate

    block_sequence_positions = _repeated_block_duplicate_positions(
        useful_fingerprints
    )
    block_removals = {
        useful_entries[position][0]
        for position in block_sequence_positions
    }

    header_removals: set[int] = set()
    header = _recognized_header_fingerprint(useful_fingerprints)
    if header is not None:
        header_positions = [
            original_position
            for original_position, fingerprint in useful_entries
            if fingerprint == header
        ]
        header_removals.update(header_positions[1:])
        header_removals.difference_update(block_removals)

    frequency_removals: set[int] = set()
    initial_counts = Counter(useful_fingerprints)
    initial_duplicate_rows = sum(
        count - 1 for count in initial_counts.values() if count > 1
    )
    initial_duplicate_ratio = (
        initial_duplicate_rows / len(useful_fingerprints)
    )
    explosion_factor = (
        len(useful_fingerprints) / max(1, len(initial_counts))
    )
    repeated_fingerprints = sum(
        count >= PDF_XLSX_FREQUENCY_MIN_OCCURRENCES
        for count in initial_counts.values()
    )
    structural_removals = block_removals | header_removals
    if (
        initial_duplicate_ratio >= PDF_XLSX_MASS_DUPLICATION_RATIO
        and explosion_factor >= PDF_XLSX_FREQUENCY_EXPLOSION_FACTOR
        and repeated_fingerprints
        >= PDF_XLSX_FREQUENCY_MIN_FINGERPRINTS
    ):
        seen: set[Tuple[str, ...]] = set()
        for original_position, fingerprint in useful_entries:
            if original_position in structural_removals:
                continue
            if fingerprint in seen:
                frequency_removals.add(original_position)
            else:
                seen.add(fingerprint)

    all_removals = (
        block_removals
        | header_removals
        | frequency_removals
    )
    if not all_removals:
        return candidate

    keep_positions = [
        position
        for position in range(len(fingerprints))
        if position not in all_removals
    ]
    cleaned = candidate.dataframe.iloc[keep_positions].copy().reset_index(drop=True)
    result = replace(
        candidate,
        dataframe=cleaned,
        duplicate_rows_removed=len(all_removals),
        repeated_block_rows_removed=len(block_removals),
        repeated_header_rows_removed=len(header_removals),
        frequency_rows_removed=len(frequency_removals),
    )
    original_score = candidate.score
    for key, value in _table_candidate_metrics(cleaned).items():
        setattr(result, key, value)
    # O score registrado é o usado na escolha, antes da limpeza do vencedor.
    result.score = original_score
    return result


def _select_and_deduplicate_candidates(
    candidates: Sequence[TableCandidate],
    *,
    emit_logs: bool = True,
) -> List[TableCandidate]:
    selected = []
    for winner in _select_candidate_winners(candidates):
        rows_before = winner.useful_rows
        duplicate_ratio = winner.duplicate_ratio
        cleaned = _deduplicate_candidate_rows(winner)
        selected.append(cleaned)
        if emit_logs:
            logger.info(
                "PDF->XLSX candidato selecionado page=%d group=%s "
                "extractor=%s rows_before=%d rows_after=%d "
                "distinct_rows=%d duplicate_ratio=%.4f score=%.2f "
                "removed_block=%d removed_header=%d removed_frequency=%d",
                winner.page_number,
                winner.group_id or "-",
                winner.extractor,
                rows_before,
                cleaned.useful_rows,
                cleaned.unique_rows,
                duplicate_ratio,
                winner.score,
                cleaned.repeated_block_rows_removed,
                cleaned.repeated_header_rows_removed,
                cleaned.frequency_rows_removed,
            )
    return selected


def _candidate_fallback_needed(
    candidates: Sequence[TableCandidate],
) -> bool:
    winners = _select_candidate_winners(candidates)
    if not winners:
        return True
    return any(
        candidate.score < PDF_XLSX_MIN_CANDIDATE_SCORE
        or candidate.duplicate_ratio >= PDF_XLSX_MASS_DUPLICATION_RATIO
        for candidate in winners
    )


def _validate_selected_pdf_xlsx_candidates(
    candidates: Sequence[TableCandidate],
) -> None:
    if not candidates or not any(candidate.useful_rows for candidate in candidates):
        raise BadRequest(
            "Não foi possível extrair linhas úteis deste PDF."
        )
    for candidate in candidates:
        if candidate.unique_rows > PDF_XLSX_MAX_UNIQUE_ROWS_PER_CANDIDATE:
            raise BadRequest(
                "A tabela extraída excede o limite seguro de linhas úteis."
            )
        if candidate.useful_columns <= 0:
            raise BadRequest(
                "A tabela extraída não possui colunas úteis."
            )
        if candidate.duplicate_ratio >= PDF_XLSX_MASS_DUPLICATION_RATIO:
            raise BadRequest(
                "A tabela extraída permaneceu massivamente duplicada."
            )
        original_rows = candidate.useful_rows + candidate.duplicate_rows_removed
        removed_ratio = (
            candidate.duplicate_rows_removed / original_rows
            if original_rows
            else 0.0
        )
        if removed_ratio >= PDF_XLSX_MASS_DUPLICATION_RATIO:
            logger.warning(
                "PDF->XLSX duplicacao massiva corrigida page=%d group=%s "
                "extractor=%s rows_before=%d rows_after=%d ratio=%.4f",
                candidate.page_number,
                candidate.group_id or "-",
                candidate.extractor,
                original_rows,
                candidate.useful_rows,
                removed_ratio,
            )
        empty_column_ratio = (
            candidate.fully_empty_columns / candidate.raw_columns
            if candidate.raw_columns
            else 1.0
        )
        if empty_column_ratio >= 0.50:
            logger.warning(
                "PDF->XLSX excesso de colunas vazias page=%d group=%s "
                "extractor=%s columns=%d empty_columns=%d",
                candidate.page_number,
                candidate.group_id or "-",
                candidate.extractor,
                candidate.raw_columns,
                candidate.fully_empty_columns,
            )
        if candidate.score < PDF_XLSX_MIN_CANDIDATE_SCORE:
            logger.warning(
                "PDF->XLSX candidato de baixa confianca page=%d group=%s "
                "extractor=%s rows=%d columns=%d score=%.2f",
                candidate.page_number,
                candidate.group_id or "-",
                candidate.extractor,
                candidate.useful_rows,
                candidate.useful_columns,
                candidate.score,
            )


def _make_unique_columns(cols: List[str]) -> List[str]:
    seen: Dict[str,int] = {}
    out: List[str] = []
    for c in cols:
        base = (c or "Coluna").strip() or "Coluna"
        if base not in seen:
            seen[base] = 1; out.append(base)
        else:
            seen[base] += 1; out.append(f"{base} ({seen[base]})")
    return out

def _trim_headers(headers: List[str], max_len: int = 80) -> List[str]:
    out: List[str] = []
    used: set[str] = set()
    for h in headers:
        s = str(h or "").strip()
        if len(s) > max_len:
            s = s[:max_len-1] + "…"
        base = s or "Coluna"
        cand, i = base, 2
        while cand in used:
            cand = f"{base} ({i})"; i += 1
        used.add(cand); out.append(cand)
    return out

def _clean_and_infer(df):
    import pandas as pd
    # Guard: df inválido ou vazio
    if df is None:
        return pd.DataFrame(), {}
    try:
        df = df.copy().map(lambda x: "" if x is None else str(x).replace("\r"," ").replace("\n"," ").strip())
    except Exception:
        return pd.DataFrame(), {}
    df = df.loc[:, (df != "").any(axis=0)]
    df = df[(df != "").any(axis=1)]
    if df.empty or len(df) == 0:
        return df, {}

    # Rastreia POSIÇÃO inteira (pos) e label (i) separadamente.
    # df.iloc usa posição; df.loc usa label — misturar os dois causa IndexError.
    header_pos, best_fill = 0, -1.0
    for pos, (i, row) in enumerate(df.iterrows()):
        non_empty = (row != "").sum()
        fill = non_empty / max(1, len(row))
        if fill > best_fill and non_empty >= 2:
            best_fill, header_pos = fill, pos
        if fill >= 0.7:
            header_pos = pos
            break

    # Valida que header_pos está dentro dos limites
    if header_pos >= len(df):
        return pd.DataFrame(), {}

    header = [h if h else f"Coluna {j+1}" for j, h in enumerate(list(df.iloc[header_pos].values))]
    header = _trim_headers(_make_unique_columns([str(h) for h in header]), max_len=80)
    df = df.iloc[header_pos + 1:].reset_index(drop=True)
    df.columns = header

    df = df[~(df.apply(lambda r: (list(r.values) == header), axis=1))]

    meta: Dict[str, Dict[str, Any]] = {}
    for col in list(df.columns):
        series = df[col].astype(str)
        parsed = [_maybe_number(v) for v in series]
        ratio = sum(p is not None for p in parsed) / max(1, len(parsed))
        if ratio >= 0.6:
            has_percent  = any(str(v).strip().endswith("%") for v in series)
            has_currency = any("R$" in str(v) for v in series)
            df[col] = [(p if p is not None else None) for p in parsed]
            meta[col] = {"type": "percent" if has_percent else ("money" if has_currency else "number")}
        else:
            df[col] = [_excel_safe_str(v) for v in series]
            meta[col] = {"type": "text"}

    df = df.loc[:, df.notna().any(axis=0)]
    if len(set(df.columns)) != len(df.columns):
        df.columns = _make_unique_columns(list(df.columns))
    return df, meta

def _norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return (s.replace("ç","c").replace("á","a").replace("à","a").replace("ã","a").replace("â","a")
              .replace("é","e").replace("ê","e").replace("í","i").replace("ó","o").replace("ô","o")
              .replace("õ","o").replace("ú","u").replace("%"," pct ").replace("º",""))

# ---------------- Alvos/schema, detecção de áreas, etc. (mantidos) ----------------
def _load_target_schema_from_env() -> Optional[List[str]]:
    import pandas as pd
    schema_file = os.environ.get("XLSM_SCHEMA_FILE")
    if not schema_file:
        default_schema = os.path.join(os.getcwd(), "envs", "modelo.xlsx")
        if os.path.exists(default_schema):
            schema_file = default_schema
    if schema_file and ((schema_file.startswith('"') and schema_file.endswith('"')) or
                        (schema_file.startswith("'") and schema_file.endswith("'"))):
        schema_file = schema_file[1:-1]
    if schema_file and os.path.exists(schema_file):
        try:
            xls = pd.ExcelFile(schema_file)
            df = pd.read_excel(schema_file, sheet_name=xls.sheet_names[0], header=None)
            for _, row in df.iterrows():
                vals = [str(v).strip() for v in row.values if str(v).strip()]
                if len(vals) >= 2:
                    return vals
        except Exception as exc:
            logger.debug(
                "PDF->XLSX schema stage=failed error=%s",
                type(exc).__name__,
            )
    cols_env = os.environ.get("XLSM_TARGET_COLUMNS")
    if cols_env:
        cols = [c.strip() for c in cols_env.split(",") if c.strip()]
        if len(cols) >= 2:
            return cols
    return None

def _bbox_plumber_to_camelot(page_height: float, bbox_plumber: Tuple[float,float,float,float]) -> str:
    x0, top, x1, bottom = bbox_plumber
    y_top_c = page_height - top
    y_bot_c = page_height - bottom
    return f"{x0},{y_top_c},{x1},{y_bot_c}"

def _cluster_positions(vals, tol: float = 2.5):
    vals = sorted(float(v) for v in vals)
    if not vals: return []
    clusters, cur = [], [vals[0]]
    for v in vals[1:]:
        if abs(v - cur[-1]) <= tol:
            cur.append(v)
        else:
            clusters.append(sum(cur)/len(cur)); cur = [v]
    clusters.append(sum(cur)/len(cur))
    return clusters

def _detect_table_bbox_and_columns(page, header_hints=None) -> Tuple[Tuple[float,float,float,float], List[float]]:
    header_hints = header_hints or ["Nome","Segurado","Valor","Part."]
    words = page.extract_words() or []
    W, H = page.width, page.height
    header_y_top, left_edge, right_edge = None, W*0.07, W*0.98
    for w in words:
        txt = (w.get("text") or "").strip()
        if any(h.lower() in txt.lower() for h in header_hints):
            if header_y_top is None or w["top"] < header_y_top:
                header_y_top = w["top"]
            left_edge = min(left_edge, w["x0"])
            right_edge = max(right_edge, w["x1"])
    if header_y_top is None:
        header_y_top = H*0.20
    bottom_edge = H*0.90
    try:
        hlines = [l for l in (page.lines or []) if abs(l["y0"]-l["y1"]) < 0.5 and (l["x1"]-l["x0"]) > (W*0.6)]
        if hlines:
            bottom_edge = max(l["y0"] for l in hlines if l["y0"] > header_y_top+5)
    except Exception:
        pass
    top, bottom = max(0, header_y_top-6), min(H-1, bottom_edge+6)
    x0, x1 = max(0, left_edge-6), min(W-1, right_edge+6)

    vxs = []
    try:
        for l in (page.lines or []):
            if abs(l["x0"]-l["x1"]) >= 0.5:  # queremos linhas verticais
                continue
            y_top, y_bot = min(l["y0"],l["y1"]), max(l["y0"],l["y1"])
            if y_bot>=top and y_top<=bottom and x0<=l["x0"]<=x1:
                vxs.append(l["x0"])
    except Exception:
        pass
    cols = [c for c in _cluster_positions(vxs, tol=2.0) if (c-x0)>5 and (x1-c)>5]
    return (x0, top, x1, bottom), cols

def _extract_tables_smart(src_pdf: str) -> List[TableCandidate]:
    import pdfplumber
    dpi = int(os.environ.get("PDF_TO_XLSX_DPI","200"))
    line_scale = int(os.environ.get("PDF_TO_XLSX_LINE_SCALE","80"))
    process_bg = os.environ.get("PDF_PROCESS_BACKGROUND","0") == "1"
    allow_stream = os.environ.get("PDF_TO_XLSX_ALLOW_STREAM","0") == "1"
    candidates: List[TableCandidate] = []

    pages_env = os.environ.get("PDF_PAGE_RANGE")
    def page_allowed(i: int) -> bool:
        if not pages_env: return True
        def parse_range(spec: str):
            for token in spec.split(','):
                token = token.strip()
                if '-' in token:
                    a,b = token.split('-',1)
                    try: a=int(a); b=int(b)
                    except: continue
                    for v in range(a,b+1): yield v
                else:
                    try: yield int(token)
                    except: pass
        return i in set(parse_range(pages_env))

    header_hints = os.environ.get(
        "PDF_HEADER_HINTS",
        "Cia,Suc,Apol.,Cob,Fatura,Estipulante,CPF,Serviço,Quantidade,Valor,Conta,Ramo,Data Emissão"
    ).split(",")

    with pdfplumber.open(src_pdf) as pdf:
        for idx, p in enumerate(pdf.pages, start=1):
            check_converter_deadline("pdf-xlsx-smart-page")
            if not page_allowed(idx): continue

            try:
                preview = (p.extract_text() or "").strip().upper()[:200]
                if "MENSAGENS" in preview:
                    continue
            except Exception:
                pass

            bbox_pl, cols = _detect_table_bbox_and_columns(p, header_hints=header_hints)
            area = _bbox_plumber_to_camelot(p.height, bbox_pl)

            found = False
            try:
                extracted = _run_camelot_worker(
                    src_pdf,
                    flavor="lattice",
                    pages=str(idx),
                    page_hint=idx,
                    extractor="camelot-lattice-smart",
                    region_prefix=f"p{idx:04d}-smart",
                    region_index_width=3,
                    table_area=area,
                    line_scale=line_scale,
                    strip_text="\n",
                    process_background=process_bg,
                    copy_text=["h", "v"],
                    shift_text=["l", "t"],
                    dpi=dpi,
                )
                candidates.extend(extracted)
                found = bool(extracted)
            except _ConverterDeadlineReservedError:
                raise
            except Exception as exc:
                logger.debug(
                    "PDF->XLSX smart_lattice stage=failed page=%d "
                    "error=%s",
                    idx,
                    type(exc).__name__,
                )

            if (not found) and allow_stream:
                check_converter_deadline("pdf-xlsx-smart-stream")
                try:
                    col_str = ",".join(str(int(x)) for x in cols) if cols else None
                    candidates.extend(_run_camelot_worker(
                        src_pdf,
                        flavor="stream",
                        pages=str(idx),
                        page_hint=idx,
                        extractor="camelot-stream-smart",
                        region_prefix=f"p{idx:04d}-smart-stream",
                        region_index_width=3,
                        table_area=area,
                        columns=col_str,
                        strip_text="\n",
                        dpi=dpi,
                    ))
                except _ConverterDeadlineReservedError:
                    raise
                except Exception as exc:
                    logger.debug(
                        "PDF->XLSX smart_stream stage=failed page=%d "
                        "error=%s",
                        idx,
                        type(exc).__name__,
                    )
    return candidates

def _find_dense_table_areas(pdf_path: str) -> Dict[int, List[str]]:
    import pdfplumber
    areas_by_page: Dict[int, List[str]] = {}
    with pdfplumber.open(pdf_path) as pdf:
        for i, p in enumerate(pdf.pages, start=1):
            check_converter_deadline("pdf-xlsx-dense-area")
            W, H = p.width, p.height
            lines, rects = p.lines or [], p.rects or []
            candidates = []
            for r in rects:
                w = abs(r["x1"] - r["x0"]); h = abs(r["y1"] - r["y0"])
                if w*h < (W*H*0.05):
                    continue
                inside = 0
                for ln in lines:
                    x0, x1 = min(ln["x0"],ln["x1"]), max(ln["x0"],ln["x1"])
                    y0, y1 = min(ln["y0"],ln["y1"]), max(ln["y0"],ln["y1"])
                    if (x0 >= r["x0"]-2 and x1 <= r["x1"]+2 and y0 >= r["y0"]-2 and y1 <= r["y1"]+2):
                        inside += 1
                score = inside / max(1.0, (w*h)/(W*H))
                candidates.append((score, r))
            candidates.sort(key=lambda t: t[0], reverse=True)
            picks = []
            for _, r in candidates[:2]:
                picks.append(_bbox_plumber_to_camelot(H, (r["x0"],r["y0"],r["x1"],r["y1"])))
            if picks:
                areas_by_page[i] = picks
    return areas_by_page

# ---------------- Escrita XLSX ----------------
def _write_minimal_xlsx(out_path: str, message: str = "Nenhuma tabela detectada. Tente habilitar OCR.") -> None:
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = "Dados"
    ws["A1"] = "Aviso"
    ws["A2"] = _neutralize_spreadsheet_formula(message)
    ws.column_dimensions["A"].width = min(80, max(20, int(len(message) * 0.9)))
    wb.save(out_path)

def _format_openpyxl_sheet(ws, col_meta: Dict[str, Dict[str, Any]]):
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    max_col, max_row = ws.max_column, ws.max_row
    if max_col and max_row:
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    for j, col_name_cell in enumerate(ws[1], start=1):
        col_name = str(col_name_cell.value or "")
        meta = col_meta.get(col_name, {"type":"text"})
        col_letter = get_column_letter(j)
        if meta["type"] == "percent":
            for r in range(2, max_row+1):
                c = ws[f"{col_letter}{r}"]; c.number_format = "0.00%"; c.alignment = Alignment(horizontal="right")
        elif meta["type"] == "money":
            for r in range(2, max_row+1):
                c = ws[f"{col_letter}{r}"]; c.number_format = "#,##0.00"; c.alignment = Alignment(horizontal="right")
        elif meta["type"] == "number":
            for r in range(2, max_row+1):
                c = ws[f"{col_letter}{r}"]; c.number_format = "#,##0.########"; c.alignment = Alignment(horizontal="right")
        else:
            if "data" in _norm(col_name):
                for r in range(2, max_row+1):
                    c = ws[f"{col_letter}{r}"]; c.number_format = "dd/mm/yyyy"; c.alignment = Alignment(horizontal="center")
            else:
                for r in range(2, max_row+1):
                    ws[f"{col_letter}{r}"].alignment = Alignment(horizontal="left")

    for j in range(1, max_col+1):
        col_letter = get_column_letter(j); max_len = 0
        for r in range(1, max_row+1):
            v = ws[f"{col_letter}{r}"].value
            if v is None: continue
            s = str(v) + ("   " if r == 1 else "")
            max_len = max(max_len, len(s))
        ws.column_dimensions[col_letter].width = max(10, min(60, int(max_len*1.15)))

    add_table = os.environ.get("XLSX_ADD_TABLE","0") == "1"
    if add_table and max_row >= 2 and max_col >= 1:
        import uuid as _uuid
        tbl_name = f"Tbl_{_uuid.uuid4().hex[:8]}"
        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        table = Table(displayName=tbl_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2",
                                             showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border

def _drop_all_empty_rows(df):
    import pandas as pd
    df2 = df.copy()
    for c in df2.columns:
        df2[c] = df2[c].replace("", pd.NA)
    df2 = df2.dropna(how="all")
    return df2

def _map_columns_to_schema_with_stats(df, target_cols: List[str]):
    import pandas as pd
    from difflib import SequenceMatcher
    src_cols = list(df.columns)
    used = set(); matched = 0

    def best_match(target: str) -> Optional[str]:
        tnorm = _norm(target)
        best, best_score = None, 0.0
        for c in src_cols:
            if c in used: continue
            score = SequenceMatcher(None, _norm(c), tnorm).ratio()
            if _norm(c) in tnorm or tnorm in _norm(c):
                score += 0.15
            if score > best_score:
                best, best_score = c, score
        return best if best_score >= 0.55 else None

    out = {}
    for tgt in target_cols:
        match = best_match(tgt)
        out[tgt] = df[match] if match else pd.Series([""] * len(df))
        if match:
            used.add(match); matched += 1
    return pd.DataFrame(out), matched, len(target_cols)

def _rescue_with_stream(src_pdf: str, pages: str) -> List[TableCandidate]:
    check_converter_deadline("pdf-xlsx-stream")
    dpi = int(os.environ.get("PDF_TO_XLSX_DPI","200"))
    resolved_pages = _camelot_worker.parse_pages_spec(pages)
    return _run_camelot_worker(
        src_pdf,
        flavor="stream",
        pages=pages,
        page_hint=min(resolved_pages),
        extractor="camelot-stream-global",
        region_prefix="stream-global",
        region_index_width=4,
        strip_text="\n",
        dpi=dpi,
    )

def _pdfplumber_tables_dfs(src_pdf: str, pages: str) -> List[TableCandidate]:
    import pdfplumber, pandas as pd
    candidates: List[TableCandidate] = []
    pages_set = None
    if pages and pages != "all":
        def parse_range(spec: str):
            for token in spec.split(','):
                token = token.strip()
                if '-' in token:
                    a,b = token.split('-',1)
                    try: a=int(a); b=int(b)
                    except: continue
                    for v in range(a,b+1): yield v
                else:
                    try: yield int(token)
                    except: pass
        pages_set = set(parse_range(pages))
    with pdfplumber.open(src_pdf) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            check_converter_deadline("pdf-xlsx-pdfplumber")
            if pages_set and i not in pages_set: continue
            found_tables = page.find_tables() or []
            if found_tables:
                for table_index, table in enumerate(found_tables, start=1):
                    data = table.extract()
                    dataframe = pd.DataFrame(data)
                    if dataframe.empty:
                        continue
                    bbox_plumber = getattr(table, "bbox", None)
                    bbox = None
                    if bbox_plumber and len(bbox_plumber) == 4:
                        x0, top, x1, bottom = bbox_plumber
                        bbox = _normalize_candidate_bbox(
                            (x0, page.height - bottom, x1, page.height - top)
                        )
                    candidates.append(_make_table_candidate(
                        dataframe,
                        page_number=i,
                        extractor="pdfplumber",
                        bbox=bbox,
                        region_id=f"p{i:04d}-pdfplumber-{table_index:03d}",
                        scope="region" if bbox is not None else "page",
                    ))
                continue

            # Compatibilidade com páginas nas quais find_tables não oferece bbox.
            for table_index, data in enumerate(
                page.extract_tables() or [],
                start=1,
            ):
                dataframe = pd.DataFrame(data)
                if dataframe.empty:
                    continue
                candidates.append(_make_table_candidate(
                    dataframe,
                    page_number=i,
                    extractor="pdfplumber",
                    region_id=f"p{i:04d}-pdfplumber-{table_index:03d}",
                    scope="page",
                ))
    return candidates


def _resolve_requested_pdf_pages(
    total_pages: int,
    pages_spec: Optional[str],
) -> List[int]:
    """Resolve a configuração de páginas uma vez, sem confiar nos candidatos."""
    total = max(0, int(total_pages or 0))
    if total <= 0:
        return []
    spec = (pages_spec or "all").strip().lower()
    if not spec or spec == "all":
        return list(range(1, total + 1))

    resolved: set[int] = set()
    for token in spec.split(","):
        token = token.strip()
        if not token:
            continue
        if "-" in token:
            first_raw, last_raw = token.split("-", 1)
            try:
                first = int(first_raw)
                last = int(last_raw)
            except ValueError:
                continue
            if last < first:
                continue
            resolved.update(
                page
                for page in range(first, last + 1)
                if 1 <= page <= total
            )
            continue
        try:
            page = int(token)
        except ValueError:
            continue
        if 1 <= page <= total:
            resolved.add(page)
    return sorted(resolved)


def _pages_to_extractor_spec(pages: Sequence[int]) -> str:
    return ",".join(str(page) for page in sorted(set(pages)))


def _fallback_reasons_by_page(
    expected_pages: Sequence[int],
    candidates: Sequence[TableCandidate],
) -> Dict[int, str]:
    winners = _select_candidate_winners(candidates)
    winners_by_page: Dict[int, List[TableCandidate]] = {}
    for candidate in winners:
        if candidate.useful_rows and candidate.useful_columns:
            winners_by_page.setdefault(candidate.page_number, []).append(
                candidate
            )

    reasons: Dict[int, str] = {}
    for page in expected_pages:
        page_winners = winners_by_page.get(page, [])
        if not page_winners:
            reasons[page] = "missing_candidate"
    return reasons


def _useful_candidates_for_pages(
    candidates: Sequence[TableCandidate],
    pages: Sequence[int],
) -> List[TableCandidate]:
    allowed = set(pages)
    return [
        candidate
        for candidate in candidates
        if (
            candidate.page_number in allowed
            and candidate.useful_rows > 0
            and candidate.useful_columns > 0
        )
    ]


def _run_pdf_xlsx_fallback_pass(
    src_pdf: str,
    expected_pages: Sequence[int],
    candidates: Sequence[TableCandidate],
    *,
    allow_stream: bool,
) -> List[TableCandidate]:
    """
    Executa no máximo uma passagem por estratégia e somente nas páginas que
    estão ausentes ou têm vencedor de baixa qualidade.
    """
    recovered: List[TableCandidate] = []
    reasons = _fallback_reasons_by_page(expected_pages, candidates)
    if not reasons:
        return recovered

    for page, reason in sorted(reasons.items()):
        check_converter_deadline("pdf-xlsx-fallback-page")
        logger.info(
            "PDF->XLSX fallback stage=start page=%d group=- "
            "extractor=- reason=%s candidate_useful=%s",
            page,
            reason,
            str(reason != "missing_candidate").lower(),
        )

    if allow_stream:
        stream_pages = sorted(reasons)
        check_converter_deadline("pdf-xlsx-fallback-stream")
        try:
            stream_candidates = _rescue_with_stream(
                src_pdf,
                _pages_to_extractor_spec(stream_pages),
            )
            recovered.extend(
                _useful_candidates_for_pages(
                    stream_candidates,
                    stream_pages,
                )
            )
        except _ConverterDeadlineReservedError:
            raise
        except Exception as exc:
            logger.debug(
                "PDF->XLSX fallback stream falhou: %s",
                type(exc).__name__,
            )
        check_converter_deadline("pdf-xlsx-fallback-stream-output")

    combined = list(candidates) + recovered
    remaining_reasons = _fallback_reasons_by_page(
        expected_pages,
        combined,
    )
    if remaining_reasons:
        plumber_pages = sorted(remaining_reasons)
        check_converter_deadline("pdf-xlsx-fallback-pdfplumber")
        try:
            plumber_candidates = _pdfplumber_tables_dfs(
                src_pdf,
                _pages_to_extractor_spec(plumber_pages),
            )
            recovered.extend(
                _useful_candidates_for_pages(
                    plumber_candidates,
                    plumber_pages,
                )
            )
        except Exception as exc:
            logger.debug(
                "PDF->XLSX fallback pdfplumber falhou: %s",
                type(exc).__name__,
            )
        check_converter_deadline("pdf-xlsx-fallback-pdfplumber-output")

    final_candidates = list(candidates) + recovered
    unresolved = _fallback_reasons_by_page(
        expected_pages,
        final_candidates,
    )
    represented = {
        candidate.page_number
        for candidate in _select_candidate_winners(final_candidates)
        if candidate.useful_rows and candidate.useful_columns
    }
    for page, initial_reason in sorted(reasons.items()):
        check_converter_deadline("pdf-xlsx-fallback-result-page")
        logger.info(
            "PDF->XLSX fallback stage=result page=%d group=- "
            "extractor=controlled_pass reason=%s candidate_useful=%s "
            "fallback_remaining=%s",
            page,
            initial_reason,
            str(page in represented).lower(),
            unresolved.get(page, "none"),
        )
    return recovered


def _structured_text_fallback_candidates(
    src_pdf: str,
    pages: Sequence[int],
) -> List[TableCandidate]:
    """
    Último fallback somente para texto com separação tabular consistente.

    Texto corrido ou página vazia não vira artificialmente uma tabela de uma
    coluna.
    """
    import pandas as pd
    import pdfplumber

    requested = set(pages)
    candidates: List[TableCandidate] = []
    with pdfplumber.open(src_pdf) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            check_converter_deadline("pdf-xlsx-structured-text-page")
            if page_number not in requested:
                continue
            lines = [
                line.strip()
                for line in (page.extract_text() or "").splitlines()
                if line.strip()
            ]
            split_rows: List[List[str]] = []
            for line in lines:
                if "\t" in line:
                    cells = [cell.strip() for cell in line.split("\t")]
                elif "|" in line:
                    cells = [cell.strip() for cell in line.split("|")]
                elif ";" in line:
                    cells = [cell.strip() for cell in line.split(";")]
                else:
                    cells = [
                        cell.strip()
                        for cell in re.split(r"\s{2,}", line)
                    ]
                cells = [cell for cell in cells if cell]
                if len(cells) >= 2:
                    split_rows.append(cells)

            column_counts = Counter(len(row) for row in split_rows)
            if not column_counts:
                continue
            modal_columns, modal_rows = max(
                column_counts.items(),
                key=lambda item: (item[1], item[0]),
            )
            if (
                modal_columns < 2
                or modal_rows < 3
                or modal_rows / max(1, len(lines)) < 0.60
            ):
                continue
            consistent_rows = [
                row for row in split_rows if len(row) == modal_columns
            ]
            dataframe = pd.DataFrame(consistent_rows)
            candidate = _make_table_candidate(
                dataframe,
                page_number=page_number,
                extractor="text",
                region_id=f"p{page_number:04d}-structured-text",
                scope="page",
            )
            if candidate.useful_rows and candidate.useful_columns >= 2:
                candidates.append(candidate)
    return candidates


# ---------------- Conversores PDF ----------------
def _pdf_to_docx(in_pdf: str, out_dir: str) -> str:
    from pdf2docx import Converter
    check_converter_deadline("pdf-docx")
    os.makedirs(out_dir, exist_ok=True)
    base = os.path.splitext(os.path.basename(in_pdf))[0]
    out_path = _unique_out_path(out_dir, base, "docx")
    cv = Converter(in_pdf)
    try:
        cv.convert(out_path, start=0, end=None)
    finally:
        cv.close()
    check_converter_deadline("pdf-docx-output")
    if not os.path.exists(out_path) or os.path.getsize(out_path) == 0:
        raise RuntimeError("pdf2docx falhou ao gerar DOCX")
    return out_path

# ============================================================
# Fase 1 — Extração inteligente: tabelas de coparticipação
# ============================================================

# Código no padrão  1.01.01012  (dentro de qualquer string)
_RE_COD_ANYWHERE = re.compile(r'\d\.\d{2}\.\d{5}')
# Código no início exato de um token
_RE_COD_START    = re.compile(r'^\d\.\d{2}\.\d{5}')
# Separador para split em múltiplos blocos numa célula gigante
_RE_COD_SPLIT    = re.compile(r'(?=\d\.\d{2}\.\d{5})')

# Valores monetários: aceita "R$ 135,00", "135,00", "135.00"
_RE_MONEY_BR  = re.compile(r'R\$\s*\d{1,3}(?:\.\d{3})*,\d{2}')
_RE_MONEY_NUM = re.compile(r'\b\d{1,3}(?:[.,]\d{3})*[.,]\d{2}\b')

# Títulos de seção
_RE_SECAO = re.compile(
    r'EXEMPLOS\s+DE\s+(?:EXAMES|CO(?:PARTICIPAÇÃO|PARTICIPA[CÇ][AÃ]O))'
    r'|OUTROS\s+EXEMPLOS',
    re.IGNORECASE,
)

# Textos de observação/rodapé — devem ser descartados
_RE_OBS = re.compile(
    r'\*\s*Obs[:\s]|Com\s+exce[çc][aã]o\s+dos\s+valores'
    r'|referem-se\s+ao\s+custo|pode\s+haver\s+varia[çc][aã]o'
    r'|ANEXO\s+[IVX]+',
    re.IGNORECASE,
)

_HEADER_TOKENS = frozenset({
    'codigo', 'código', 'procedimento', 'valor unimed',
    'co-part', 'flex', 'pleno', 'cop.', 'co-participação',
    'coparticipação', '20%', '30%', '40%', '50%',
})


def _format_brl(raw: str) -> str:
    """
    Normaliza um valor monetário bruto para "R$ NNN,NN".
    Aceita: "R$ 135,00", "135,00", "135.00", "1.350,00", "1350.00".
    Retorna a string original se não conseguir converter.
    """
    s = str(raw).strip()
    # já está no formato desejado
    if _RE_MONEY_BR.match(s):
        return s
    # remove prefixo R$
    s_clean = re.sub(r'^R\$\s*', '', s).strip()
    # normaliza separadores: "1.350,00" → "1350.00", "135,00" → "135.00"
    if ',' in s_clean and '.' in s_clean:
        # formato BR com milhar: 1.350,00
        s_clean = s_clean.replace('.', '').replace(',', '.')
    elif ',' in s_clean:
        # sem milhar: 135,00
        s_clean = s_clean.replace(',', '.')
    try:
        val = float(s_clean)
        # formata de volta para BR
        # separa inteiro e decimal
        inteiro = int(val)
        cents   = round((val - inteiro) * 100)
        # milhar no estilo BR
        inteiro_fmt = f'{inteiro:,}'.replace(',', '.')
        return f'R$ {inteiro_fmt},{cents:02d}'
    except (ValueError, TypeError):
        return raw  # devolve original se falhar


def _extract_money_values(text: str) -> List[str]:
    """
    Extrai todos os valores monetários de um texto, normalizados para R$ NNN,NN.
    Tenta primeiro o padrão R$; depois padrão numérico genérico.
    """
    found = _RE_MONEY_BR.findall(text)
    if found:
        return [_format_brl(v) for v in found]
    # fallback: padrão numérico (ex.: "135.00" ou "135,00")
    found_num = _RE_MONEY_NUM.findall(text)
    return [_format_brl(v) for v in found_num]


def _is_obs_line(text: str) -> bool:
    """Retorna True se a linha é rodapé/observação e deve ser descartada."""
    return bool(_RE_OBS.search(text))


def _is_header_line(text: str) -> bool:
    """Retorna True se a linha é cabeçalho de tabela a ser descartado."""
    low = text.lower()
    hits = sum(1 for tok in _HEADER_TOKENS if tok in low)
    # cabeçalho puro: ≥3 tokens E sem código de procedimento
    return hits >= 3 and not _RE_COD_ANYWHERE.search(text)


def _is_secao_line(text: str) -> bool:
    return bool(_RE_SECAO.search(text))


def _is_total_line(text: str) -> bool:
    return 'TOTAL DE COPARTICI' in text.upper()


def _row_to_text(row) -> str:
    """Une as células não-vazias de uma linha do DataFrame."""
    return ' '.join(str(v).strip() for v in row if str(v).strip())


def _df_to_lines(df) -> List[str]:
    """
    Converte cada linha do DataFrame em uma string de texto limpa,
    descartando rodapés e células vazias.
    Linhas que contenham vários códigos são divididas em sublistas.
    """
    lines: List[str] = []
    for _, row in df.iterrows():
        raw = _row_to_text(row)
        if not raw:
            continue
        if _is_obs_line(raw):
            continue
        # célula gigante com múltiplos códigos concatenados?
        parts = _RE_COD_SPLIT.split(raw)
        parts = [p.strip() for p in parts if p.strip()]
        lines.extend(parts)
    return lines


def _looks_like_coparticipacao_table(df) -> bool:
    """Retorna True se o DataFrame parece ser tabela de coparticipação."""
    if df is None or df.empty:
        return False
    text = ' '.join(
        str(v) for row in df.values for v in row if v is not None
    ).upper()
    signals = 0
    if 'CO-PART' in text:
        signals += 2
    if 'VALOR UNIMED' in text:
        signals += 2
    if 'TOTAL DE COPARTICI' in text:
        signals += 2
    if _RE_SECAO.search(text):
        signals += 2
    if _RE_COD_ANYWHERE.search(text):
        signals += 2
    return signals >= 4


def _normalize_coparticipacao_table(df) -> 'pd.DataFrame':
    """
    Normaliza um DataFrame bruto de tabela de coparticipação para o esquema
    fixo de 9 colunas. Estratégia:
      1. Converter cada linha em texto plano (via _df_to_lines).
      2. Classificar cada linha em: seção / cabeçalho / obs / total /
         procedimento / continuação.
      3. Montar records limpos, concatenando continuações ao procedimento
         anterior.
      4. Normalizar valores para "R$ NNN,NN".
      5. Limitar a 5 valores por linha de procedimento (Valor Unimed +
         4 faixas de copart.) para evitar coluna extra duplicada.
    """
    import pandas as pd

    OUT_COLS = [
        'Secao', 'Codigo', 'Procedimento',
        'Valor Unimed', 'Copart 20%', 'Copart 30%', 'Copart 40%', 'Copart 50%',
        'Tipo Linha',
    ]

    lines = _df_to_lines(df)
    rows_out: List[Dict[str, Any]] = []
    secao_atual = ''
    last_proc_idx: Optional[int] = None

    for line in lines:
        # --- descarta obs/rodapé (segunda guarda — já filtrado em _df_to_lines,
        #     mas pode aparecer concatenado)
        if _is_obs_line(line):
            continue

        # --- cabeçalho de tabela
        if _is_header_line(line):
            continue

        # --- título de seção
        if _is_secao_line(line):
            secao_atual = line.strip()
            last_proc_idx = None
            continue

        # --- linha de total
        if _is_total_line(line):
            moneys = _extract_money_values(line)
            rows_out.append({
                'Secao':        secao_atual,
                'Codigo':       '',
                'Procedimento': 'Total de Coparticipação',
                'Valor Unimed': '',
                'Copart 20%':   moneys[0] if len(moneys) > 0 else '',
                'Copart 30%':   moneys[1] if len(moneys) > 1 else '',
                'Copart 40%':   moneys[2] if len(moneys) > 2 else '',
                'Copart 50%':   moneys[3] if len(moneys) > 3 else '',
                'Tipo Linha':   'total',
            })
            last_proc_idx = None
            continue

        # --- linha de procedimento (começa com código)
        m_cod = _RE_COD_ANYWHERE.search(line)
        if m_cod and line.strip().startswith(m_cod.group()):
            cod = m_cod.group()
            rest = line[m_cod.end():].strip()
            moneys = _extract_money_values(rest)
            # isola texto do procedimento (sem os valores monetários)
            proc_text = _RE_MONEY_NUM.sub('', _RE_MONEY_BR.sub('', rest)).strip()
            proc_text = re.sub(r'\s{2,}', ' ', proc_text).strip()
            # limita a 5 valores: Valor Unimed + Copart 20-50%
            moneys = moneys[:5]
            rows_out.append({
                'Secao':        secao_atual,
                'Codigo':       cod,
                'Procedimento': proc_text,
                'Valor Unimed': moneys[0] if len(moneys) > 0 else '',
                'Copart 20%':   moneys[1] if len(moneys) > 1 else '',
                'Copart 30%':   moneys[2] if len(moneys) > 2 else '',
                'Copart 40%':   moneys[3] if len(moneys) > 3 else '',
                'Copart 50%':   moneys[4] if len(moneys) > 4 else '',
                'Tipo Linha':   'procedimento',
            })
            last_proc_idx = len(rows_out) - 1
            continue

        # --- possível continuação de procedimento
        moneys = _extract_money_values(line)
        if last_proc_idx is not None:
            r = rows_out[last_proc_idx]
            if not moneys:
                # puro texto: concatena ao procedimento anterior
                r['Procedimento'] = (r['Procedimento'] + ' ' + line.strip()).strip()
            else:
                # linha com valores que não chegaram na linha de código
                # (Camelot às vezes quebra assim)
                money5 = moneys[:5]
                if not r['Valor Unimed']:
                    r['Valor Unimed'] = money5[0] if money5 else ''
                    r['Copart 20%']   = money5[1] if len(money5) > 1 else r['Copart 20%']
                    r['Copart 30%']   = money5[2] if len(money5) > 2 else r['Copart 30%']
                    r['Copart 40%']   = money5[3] if len(money5) > 3 else r['Copart 40%']
                    r['Copart 50%']   = money5[4] if len(money5) > 4 else r['Copart 50%']
                else:
                    # complementa apenas colunas faltantes
                    cols_val = ['Copart 20%', 'Copart 30%', 'Copart 40%', 'Copart 50%']
                    mi = 0
                    for col in cols_val:
                        if not r[col] and mi < len(moneys):
                            r[col] = moneys[mi]
                            mi += 1
        # linha sem código e sem last_proc_idx → ignora

    if not rows_out:
        return pd.DataFrame(columns=OUT_COLS)

    result = pd.DataFrame(rows_out, columns=OUT_COLS)
    result = result[(result != '').any(axis=1)]
    return result


def _assess_post_normalization(
    before_dataframe: Any,
    after_dataframe: Any,
) -> NormalizationAssessment:
    """Compara estrutura sem alterar ou expor valores das células."""
    before = _table_candidate_metrics(before_dataframe)
    after = _table_candidate_metrics(after_dataframe)
    reasons: List[str] = []

    before_rows = max(1, before["useful_rows"])
    before_unique = max(1, before["unique_rows"])
    before_columns = max(1, before["useful_columns"])
    row_growth = after["useful_rows"] / before_rows
    unique_growth = after["unique_rows"] / before_unique

    if after["useful_rows"] <= 0 or after["useful_columns"] <= 0:
        reasons.append("no_useful_content")

    if (
        row_growth >= PDF_XLSX_POST_NORMALIZATION_ROW_GROWTH
        and after["duplicate_ratio"]
        >= PDF_XLSX_POST_NORMALIZATION_DUPLICATE_RATIO
        and unique_growth
        < row_growth * PDF_XLSX_POST_NORMALIZATION_UNIQUE_SHARE
    ):
        reasons.append("row_growth_without_information")

    if (
        after["duplicate_ratio"]
        >= PDF_XLSX_POST_NORMALIZATION_DUPLICATE_RATIO
        and (
            after["duplicate_ratio"] - before["duplicate_ratio"]
            >= PDF_XLSX_POST_NORMALIZATION_DUPLICATE_INCREASE
        )
        and unique_growth
        < row_growth * PDF_XLSX_POST_NORMALIZATION_UNIQUE_SHARE
    ):
        reasons.append("duplicate_growth")

    if (
        before["useful_rows"] > 0
        and after["useful_rows"] / before["useful_rows"]
        < PDF_XLSX_POST_NORMALIZATION_MIN_ROW_RETENTION
        and after["unique_rows"] / before_unique
        < PDF_XLSX_POST_NORMALIZATION_MIN_UNIQUE_RETENTION
    ):
        reasons.append("unjustified_row_loss")

    if (
        after["near_empty_ratio"]
        >= PDF_XLSX_POST_NORMALIZATION_NEAR_EMPTY_RATIO
        and (
            after["near_empty_ratio"] - before["near_empty_ratio"]
            >= PDF_XLSX_POST_NORMALIZATION_NEAR_EMPTY_INCREASE
        )
    ):
        reasons.append("near_empty_growth")

    if (
        after["useful_columns"]
        >= before_columns * PDF_XLSX_POST_NORMALIZATION_COLUMN_GROWTH
        and after["useful_columns"]
        >= (
            before_columns
            + PDF_XLSX_POST_NORMALIZATION_COLUMN_ALLOWANCE
        )
    ):
        reasons.append("column_explosion")

    if (
        before["density"] > 0
        and after["density"]
        < PDF_XLSX_POST_NORMALIZATION_MIN_DENSITY
        and after["density"]
        < (
            before["density"]
            * PDF_XLSX_POST_NORMALIZATION_MIN_DENSITY_RETENTION
        )
    ):
        reasons.append("density_loss")

    if after["unique_rows"] > PDF_XLSX_MAX_UNIQUE_ROWS_PER_CANDIDATE:
        reasons.append("unique_row_limit")

    if (
        after["duplicate_ratio"] >= PDF_XLSX_MASS_DUPLICATION_RATIO
        and row_growth >= PDF_XLSX_POST_NORMALIZATION_ROW_GROWTH
    ):
        reasons.append("mass_duplication")

    return NormalizationAssessment(
        accepted=not reasons,
        reasons=tuple(dict.fromkeys(reasons)),
        before=before,
        after=after,
    )


def _dataframe_is_structurally_usable(dataframe: Any) -> bool:
    metrics = _table_candidate_metrics(dataframe)
    return (
        metrics["useful_rows"] > 0
        and metrics["useful_columns"] > 0
        and metrics["unique_rows"]
        <= PDF_XLSX_MAX_UNIQUE_ROWS_PER_CANDIDATE
        and metrics["duplicate_ratio"]
        < PDF_XLSX_MASS_DUPLICATION_RATIO
    )


def _validate_final_pdf_xlsx_table(
    dataframe: Any,
    *,
    page_number: int,
    group_id: str,
    extractor: str,
) -> None:
    metrics = _table_candidate_metrics(dataframe)
    if metrics["useful_rows"] <= 0 or metrics["useful_columns"] <= 0:
        raise BadRequest(
            "A tabela final não possui conteúdo útil para exportação."
        )
    if metrics["unique_rows"] > PDF_XLSX_MAX_UNIQUE_ROWS_PER_CANDIDATE:
        raise BadRequest(
            "A tabela final excede o limite seguro de linhas úteis."
        )
    if metrics["duplicate_ratio"] >= PDF_XLSX_MASS_DUPLICATION_RATIO:
        raise BadRequest(
            "A tabela final permaneceu massivamente duplicada."
        )
    if metrics["near_empty_ratio"] >= 0.90 and metrics["density"] < 0.10:
        raise BadRequest(
            "A tabela final ficou estruturalmente degradada."
        )
    logger.info(
        "PDF->XLSX final_table stage=validated page=%d group=%s "
        "extractor=%s rows=%d distinct_rows=%d duplicate_ratio=%.4f",
        page_number,
        group_id or "-",
        extractor,
        metrics["useful_rows"],
        metrics["unique_rows"],
        metrics["duplicate_ratio"],
    )


def _prepare_candidate_for_workbook(
    candidate: TableCandidate,
    *,
    sheet_name: str,
    allow_specialized_normalization: bool,
) -> PreparedWorkbookTable:
    generic, generic_meta = _clean_and_infer(candidate.dataframe)
    generic = _drop_all_empty_rows(generic)
    final_dataframe = generic
    final_meta = generic_meta
    normalization_accepted = False

    if (
        allow_specialized_normalization
        and _looks_like_coparticipacao_table(candidate.dataframe)
    ):
        try:
            normalized = _normalize_coparticipacao_table(
                candidate.dataframe
            )
        except Exception as exc:
            logger.warning(
                "PDF->XLSX normalization stage=rejected page=%d group=%s "
                "extractor=%s reason=execution_error error=%s",
                candidate.page_number,
                candidate.group_id or "-",
                candidate.extractor,
                type(exc).__name__,
            )
        else:
            assessment = _assess_post_normalization(
                candidate.dataframe,
                normalized,
            )
            logger.info(
                "PDF->XLSX normalization stage=assessed page=%d group=%s "
                "extractor=%s rows_before=%d rows_after=%d "
                "distinct_before=%d distinct_after=%d "
                "duplicate_ratio_before=%.4f duplicate_ratio_after=%.4f "
                "accepted=%s reasons=%s",
                candidate.page_number,
                candidate.group_id or "-",
                candidate.extractor,
                assessment.before["useful_rows"],
                assessment.after["useful_rows"],
                assessment.before["unique_rows"],
                assessment.after["unique_rows"],
                assessment.before["duplicate_ratio"],
                assessment.after["duplicate_ratio"],
                str(assessment.accepted).lower(),
                ",".join(assessment.reasons) or "none",
            )
            if (
                assessment.accepted
                and _dataframe_is_structurally_usable(normalized)
            ):
                final_dataframe = normalized
                final_meta = {}
                normalization_accepted = True
            else:
                logger.warning(
                    "PDF->XLSX normalization stage=fallback page=%d "
                    "group=%s extractor=%s fallback=pre_normalization",
                    candidate.page_number,
                    candidate.group_id or "-",
                    candidate.extractor,
                )

    if not _dataframe_is_structurally_usable(final_dataframe):
        raise BadRequest(
            "A normalização não gerou uma tabela estruturalmente segura."
        )
    _validate_final_pdf_xlsx_table(
        final_dataframe,
        page_number=candidate.page_number,
        group_id=candidate.group_id,
        extractor=candidate.extractor,
    )
    return PreparedWorkbookTable(
        dataframe=final_dataframe,
        meta=final_meta,
        sheet_name=sheet_name,
        page_number=candidate.page_number,
        group_id=candidate.group_id,
        extractor=candidate.extractor,
        normalization_accepted=normalization_accepted,
    )


def _validate_pdf_xlsx_consolidation(
    source_dataframes: Sequence[Any],
    consolidated: Any,
) -> None:
    source_metrics = [
        _table_candidate_metrics(dataframe)
        for dataframe in source_dataframes
    ]
    final_metrics = _table_candidate_metrics(consolidated)
    source_rows = sum(metrics["useful_rows"] for metrics in source_metrics)
    source_columns = max(
        (metrics["useful_columns"] for metrics in source_metrics),
        default=0,
    )

    if final_metrics["useful_rows"] <= 0 or final_metrics["useful_columns"] <= 0:
        raise BadRequest(
            "A consolidação não possui conteúdo útil para exportação."
        )
    if (
        source_rows > 0
        and final_metrics["useful_rows"] > max(10, int(source_rows * 1.10))
    ):
        raise BadRequest(
            "A consolidação multiplicou linhas sem origem correspondente."
        )
    if (
        source_columns > 0
        and final_metrics["useful_columns"]
        >= source_columns * PDF_XLSX_POST_NORMALIZATION_COLUMN_GROWTH
        and final_metrics["useful_columns"]
        >= source_columns + PDF_XLSX_POST_NORMALIZATION_COLUMN_ALLOWANCE
    ):
        raise BadRequest(
            "A consolidação multiplicou colunas sem origem correspondente."
        )


def _pdf_to_xlsx(in_pdf: str, out_dir: str) -> str:
    """Extrator no estilo 'modelo' ou retrocompat, controlado por env."""
    check_converter_deadline("pdf-xlsx")
    model_style = (os.environ.get("PDF_TO_XLSX_MODEL_STYLE", "0") == "1")
    page_count = enforce_pdf_page_limit(in_pdf, label="PDF de entrada")
    pages_arg = os.environ.get("PDF_PAGE_RANGE") or "all"
    expected_pages = _resolve_requested_pdf_pages(page_count, pages_arg)
    if not expected_pages:
        raise BadRequest(
            "A configuração de páginas não selecionou páginas válidas."
        )
    extractor_pages_spec = _pages_to_extractor_spec(expected_pages)
    _prepare_camelot_env()

    import pandas as pd
    t_start = time.perf_counter()
    base = os.path.splitext(os.path.basename(in_pdf))[0]
    out_dir = os.path.abspath(out_dir); os.makedirs(out_dir, exist_ok=True)
    out_path = _unique_out_path(out_dir, base, "xlsx")

    # OCR se necessário
    has_text = _pdf_has_selectable_text(in_pdf)
    check_converter_deadline("pdf-xlsx-text-detection")
    src_pdf = in_pdf if has_text else _try_ocr(in_pdf)
    check_converter_deadline("pdf-xlsx-ocr")
    geometry_by_page = _map_pdf_table_geometry_shadow(
        src_pdf,
        expected_pages,
    )

    # ---- Extração com procedência
    candidates: List[TableCandidate] = []

    # SMART BBOX (lattice + hints)
    if os.environ.get("PDF_TO_XLSX_USE_SMART_BBOX","1") == "1":
        check_converter_deadline("pdf-xlsx-smart")
        try:
            candidates.extend(_extract_tables_smart(src_pdf))
            logger.debug("SMART encontrou %d candidatos", len(candidates))
        except _ConverterDeadlineReservedError:
            raise
        except Exception as exc:
            logger.debug(
                "PDF->XLSX smart stage=failed error=%s",
                type(exc).__name__,
            )
        check_converter_deadline("pdf-xlsx-smart-output")

    # Lattice global (todas as páginas)
    if not candidates:
        check_converter_deadline("pdf-xlsx-lattice")
        try:
            dpi = int(os.environ.get("PDF_TO_XLSX_DPI","200"))
            line_scale = int(os.environ.get("PDF_TO_XLSX_LINE_SCALE","80"))
            process_bg = int(os.environ.get("PDF_PROCESS_BACKGROUND","0")) == 1
            candidates.extend(_run_camelot_worker(
                src_pdf,
                flavor="lattice",
                pages=extractor_pages_spec,
                page_hint=min(expected_pages),
                extractor="camelot-lattice-global",
                region_prefix="lattice-global",
                region_index_width=4,
                line_scale=line_scale,
                strip_text="\n",
                process_background=process_bg,
                copy_text=["h", "v"],
                shift_text=["l", "t"],
                dpi=dpi,
            ))
        except _ConverterDeadlineReservedError:
            raise
        except Exception as exc:
            logger.debug(
                "PDF->XLSX lattice_global stage=failed error=%s",
                type(exc).__name__,
            )
        check_converter_deadline("pdf-xlsx-lattice-output")

    # Lattice por áreas densas
    if (not candidates) or os.environ.get("PDF_TO_XLSX_ALWAYS_AREAS","1") == "1":
        check_converter_deadline("pdf-xlsx-areas")
        try:
            dpi = int(os.environ.get("PDF_TO_XLSX_DPI","200"))
            line_scale = int(os.environ.get("PDF_TO_XLSX_LINE_SCALE","80"))
            process_bg = os.environ.get("PDF_PROCESS_BACKGROUND","0") == "1"
            areas_by_page = _find_dense_table_areas(src_pdf)
            for page_idx, areas in areas_by_page.items():
                if page_idx not in expected_pages:
                    continue
                for area_index, area in enumerate(areas, start=1):
                    try:
                        candidates.extend(_run_camelot_worker(
                            src_pdf,
                            flavor="lattice",
                            pages=str(page_idx),
                            page_hint=page_idx,
                            extractor="camelot-lattice-region",
                            region_prefix=(
                                f"p{page_idx:04d}-dense-"
                                f"{area_index:03d}"
                            ),
                            region_index_width=3,
                            table_area=area,
                            line_scale=line_scale,
                            strip_text="\n",
                            process_background=process_bg,
                            copy_text=["h", "v"],
                            shift_text=["l", "t"],
                            dpi=dpi,
                        ))
                    except _ConverterDeadlineReservedError:
                        raise
                    except Exception as exc:
                        logger.debug(
                            "PDF->XLSX dense_area stage=failed page=%d "
                            "region=%d error=%s",
                            page_idx,
                            area_index,
                            type(exc).__name__,
                        )
        except _ConverterDeadlineReservedError:
            raise
        except Exception as exc:
            logger.debug(
                "PDF->XLSX dense_areas stage=failed error=%s",
                type(exc).__name__,
            )
        check_converter_deadline("pdf-xlsx-areas-output")

    # Uma única passagem controlada nas páginas ausentes ou de baixa qualidade.
    allow_stream_global = os.environ.get("PDF_TO_XLSX_ALLOW_STREAM","0") == "1"
    candidates.extend(
        _run_pdf_xlsx_fallback_pass(
            src_pdf,
            expected_pages,
            candidates,
            allow_stream=allow_stream_global,
        )
    )

    # Texto corrido não vira tabela artificial. Este último fallback exige
    # separação tabular consistente e só roda quando nenhum extrator retornou
    # candidato útil.
    if not candidates:
        check_converter_deadline("pdf-xlsx-structured-text-fallback")
        try:
            candidates.extend(
                _structured_text_fallback_candidates(
                    src_pdf,
                    expected_pages,
                )
            )
        except Exception as exc:
            logger.debug(
                "PDF->XLSX fallback texto estruturado falhou: %s",
                type(exc).__name__,
            )
        check_converter_deadline("pdf-xlsx-structured-text-output")

    _associate_geometry_reports(candidates, geometry_by_page)
    selected_candidates = _select_and_deduplicate_candidates(candidates)
    _validate_selected_pdf_xlsx_candidates(selected_candidates)
    tables_to_write: List[PreparedWorkbookTable] = []

    # MODEL STYLE: 1 tabela = 1 sheet ("Table N")
    if model_style:
        copart_count = 0
        for index, candidate in enumerate(selected_candidates, start=1):
            check_converter_deadline("pdf-xlsx-prepare-table")
            prepared = _prepare_candidate_for_workbook(
                candidate,
                sheet_name=f"Table {index}",
                allow_specialized_normalization=True,
            )
            if prepared.normalization_accepted:
                copart_count += 1
                prepared.sheet_name = (
                    "Coparticipacao"
                    if copart_count == 1
                    else f"Coparticipacao {index}"
                )
            tables_to_write.append(prepared)

    else:
        # Retrocompatibilidade (com consolidação e/ou schema)
        target_schema = _load_target_schema_from_env()
        single = os.environ.get("XLSX_SINGLE_SHEET","0") == "1"
        if target_schema or single:
            prepared_sources = [
                _prepare_candidate_for_workbook(
                    candidate,
                    sheet_name=f"Source {index}",
                    allow_specialized_normalization=False,
                )
                for index, candidate in enumerate(
                    selected_candidates,
                    start=1,
                )
            ]
            source_dataframes = [
                prepared.dataframe for prepared in prepared_sources
            ]
            big = (
                pd.concat(
                    source_dataframes,
                    ignore_index=True,
                    sort=False,
                )
                if source_dataframes
                else None
            )
            final_dataframe = (
                _drop_all_empty_rows(big)
                if big is not None
                else None
            )
            if (
                target_schema
                and final_dataframe is not None
                and not final_dataframe.empty
            ):
                mapped, matched, total = _map_columns_to_schema_with_stats(
                    final_dataframe,
                    target_schema,
                )
                if matched >= max(4, int(0.4 * max(1, total))):
                    final_dataframe = _drop_all_empty_rows(mapped)

            if final_dataframe is not None and not final_dataframe.empty:
                _validate_pdf_xlsx_consolidation(
                    source_dataframes,
                    final_dataframe,
                )
                _ignored, final_meta = _clean_and_infer(
                    final_dataframe.copy()
                )
                tables_to_write.append(PreparedWorkbookTable(
                    dataframe=final_dataframe,
                    meta=final_meta,
                    sheet_name="Dados",
                    page_number=0,
                    group_id="consolidated",
                    extractor="consolidation",
                ))
        else:
            copart_count = 0
            for index, candidate in enumerate(
                selected_candidates,
                start=1,
            ):
                check_converter_deadline("pdf-xlsx-prepare-table")
                prepared = _prepare_candidate_for_workbook(
                    candidate,
                    sheet_name=f"Tabela {index}",
                    allow_specialized_normalization=True,
                )
                if prepared.normalization_accepted:
                    copart_count += 1
                    prepared.sheet_name = (
                        "Coparticipacao"
                        if copart_count == 1
                        else f"Coparticipacao {index}"
                    )
                tables_to_write.append(prepared)
    if not tables_to_write:
        logger.info(
            "PDF->XLSX stage=pre_write_validation result=no_useful_tables"
        )
        raise BadRequest(
            "Não foi possível extrair tabelas úteis deste PDF."
        )

    # O workbook só começa a ser criado depois que todas as tabelas finais
    # passaram pela validação semântica interna.
    check_converter_deadline("pdf-xlsx-write")
    try:
        import xlsxwriter
        try:
            writer = pd.ExcelWriter(
                out_path,
                engine="xlsxwriter",
                engine_kwargs={
                    "options": {"strings_to_urls": False}
                },
            )
        except TypeError:
            writer = pd.ExcelWriter(out_path, engine="xlsxwriter")
    except Exception:
        writer = pd.ExcelWriter(out_path, engine="openpyxl")

    metas: Dict[str, Dict[str, Any]] = {}
    try:
        for prepared in tables_to_write:
            check_converter_deadline("pdf-xlsx-write-sheet")
            _write_dataframe_to_spreadsheet(
                prepared.dataframe,
                writer,
                sheet_name=prepared.sheet_name,
            )
            metas[prepared.sheet_name] = prepared.meta
        writer.close()
    except BaseException:
        try:
            writer.close()
        except Exception:
            pass
        try:
            if os.path.exists(out_path):
                os.remove(out_path)
        except OSError:
            pass
        raise
    check_converter_deadline("pdf-xlsx-writer-close")

    # Pós-formatação com openpyxl (larguras, filtros, números, Excel Table opcional)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(out_path)
        for ws in wb.worksheets:
            _format_openpyxl_sheet(ws, metas.get(ws.title, {}))
        wb.save(out_path)
    except Exception as exc:
        logger.debug(
            "PDF->XLSX formatting stage=failed error=%s",
            type(exc).__name__,
        )
    check_converter_deadline("pdf-xlsx-output")

    logger.info("Tempo PDF→XLSX total: %.2fs", time.perf_counter()-t_start)
    return out_path

# ---------------- PDF → CSV ----------------
FILTER_DOCX = "Office Open XML Text"
FILTER_CSV  = os.environ.get("CSV_FILTER_NAME", "Text - txt - csv (StarCalc)")
CSV_FILTER_OPTS = os.environ.get("CSV_FILTER_OPTS", "59,34,76,1")
FILTER_XLSM = "Calc MS Excel 2007 VBA XML"
FILTER_XLSX = "Calc MS Excel 2007 XML"

def _pdf_csv_cell_has_content(value: Any) -> bool:
    return value is not None and bool(str(value).strip())


def _pdf_csv_useful_dataframe(dataframe):
    if dataframe is None or getattr(dataframe, "empty", True):
        return None
    normalized = dataframe.copy().fillna("")
    content = normalized.map(_pdf_csv_cell_has_content)
    normalized = normalized.loc[content.any(axis=1), content.any(axis=0)]
    normalized = normalized.reset_index(drop=True)
    if normalized.shape[0] < 2 or normalized.shape[1] < 1:
        return None
    return normalized


def _pdf_csv_useful_rows(table) -> List[List[Any]]:
    source_rows = [list(row or []) for row in (table or [])]
    if not source_rows:
        return []
    width = max((len(row) for row in source_rows), default=0)
    if width < 1:
        return []
    padded = [row + [""] * (width - len(row)) for row in source_rows]
    padded = [row for row in padded if any(_pdf_csv_cell_has_content(v) for v in row)]
    if len(padded) < 2:
        return []
    useful_columns = [
        index
        for index in range(width)
        if any(_pdf_csv_cell_has_content(row[index]) for row in padded)
    ]
    if not useful_columns:
        return []
    return [[row[index] for index in useful_columns] for row in padded]


def _write_pdf_csv_atomically(out_path: str, write_callback) -> str:
    partial_path = os.path.join(
        os.path.dirname(out_path),
        f".{os.path.basename(out_path)}.{secrets.token_hex(8)}.partial",
    )
    try:
        write_callback(partial_path)
        if os.path.getsize(partial_path) <= 0:
            raise OSError("CSV temporário vazio.")
        os.replace(partial_path, out_path)
        return out_path
    except BaseException:
        for candidate in (partial_path, out_path):
            try:
                os.remove(candidate)
            except OSError:
                pass
        raise


def _write_pdf_csv_dataframes(out_path: str, dataframes) -> str:
    import pandas as pd

    combined = pd.concat(dataframes, ignore_index=True).fillna("")
    protected = combined.map(_neutralize_csv_field_for_spreadsheet)
    return _write_pdf_csv_atomically(
        out_path,
        lambda partial_path: protected.to_csv(
            partial_path,
            index=False,
            header=False,
            encoding="utf-8",
        ),
    )


def _write_pdf_csv_rows(out_path: str, tables) -> str:
    rows = [row for table in tables for row in table]
    width = max((len(row) for row in rows), default=0)
    protected_rows = [
        [
            _neutralize_csv_field_for_spreadsheet(value)
            for value in row + [""] * (width - len(row))
        ]
        for row in rows
    ]

    def _write(partial_path: str) -> None:
        with open(partial_path, "w", newline="", encoding="utf-8") as stream:
            csv.writer(stream).writerows(protected_rows)

    return _write_pdf_csv_atomically(out_path, _write)


def _pdf_to_csv(in_pdf: str, out_dir: str) -> str:
    check_converter_deadline("pdf-csv")
    enforce_pdf_page_limit(in_pdf, label="PDF de entrada")
    _prepare_camelot_env()
    base = os.path.splitext(os.path.basename(in_pdf))[0]
    out_dir = os.path.abspath(out_dir)
    os.makedirs(out_dir, exist_ok=True)
    out_path = _unique_out_path(out_dir, base, "csv")
    dpi = int(os.environ.get("PDF_TO_XLSX_DPI", "200"))
    line_scale = int(os.environ.get("PDF_TO_XLSX_LINE_SCALE", "80"))
    process_bg = os.environ.get("PDF_PROCESS_BACKGROUND", "0") == "1"
    pages_arg = os.environ.get("PDF_PAGE_RANGE") or "all"
    extractor_failures = 0

    for flavor in ("lattice", "stream"):
        check_converter_deadline(f"pdf-csv-{flavor}")
        useful = []
        try:
            import camelot

            options = {
                "flavor": flavor,
                "pages": pages_arg,
                "strip_text": "\n",
                "dpi": dpi,
            }
            if flavor == "lattice":
                options.update({
                    "line_scale": line_scale,
                    "process_background": process_bg,
                    "copy_text": ["h", "v"],
                    "shift_text": ["l", "t"],
                })
            collection = camelot.read_pdf(in_pdf, **options)
            check_converter_deadline(f"pdf-csv-{flavor}-output")
            for table in getattr(collection, "tables", collection):
                dataframe = _pdf_csv_useful_dataframe(getattr(table, "df", None))
                if dataframe is not None:
                    useful.append(dataframe)
        except ConverterTimeoutError:
            raise
        except Exception as exc:
            if useful:
                raise ConverterExtractionError(
                    "A extração do PDF terminou de forma incompleta."
                ) from exc
            extractor_failures += 1
            logger.debug("PDF→CSV: falha controlada no extrator %s.", flavor)
            continue

        if useful:
            try:
                return _write_pdf_csv_dataframes(out_path, useful)
            except ConverterTimeoutError:
                raise
            except Exception as exc:
                raise ConverterExtractionError(
                    "Falha ao serializar tabela extraída do PDF."
                ) from exc

    check_converter_deadline("pdf-csv-pdfplumber")
    useful_tables = []
    pdfplumber_completed = False
    try:
        import pdfplumber

        with pdfplumber.open(in_pdf) as pdf:
            for page in pdf.pages:
                check_converter_deadline("pdf-csv-pdfplumber-page")
                for table in (page.extract_tables() or []):
                    useful = _pdf_csv_useful_rows(table)
                    if useful:
                        useful_tables.append(useful)
        pdfplumber_completed = True
        check_converter_deadline("pdf-csv-pdfplumber-output")
    except ConverterTimeoutError:
        raise
    except Exception as exc:
        if useful_tables:
            raise ConverterExtractionError(
                "A extração do PDF terminou de forma incompleta."
            ) from exc
        extractor_failures += 1
        logger.debug("PDF→CSV: falha controlada no extrator de tabelas.")

    if useful_tables:
        try:
            return _write_pdf_csv_rows(out_path, useful_tables)
        except ConverterTimeoutError:
            raise
        except Exception as exc:
            raise ConverterExtractionError(
                "Falha ao serializar tabela extraída do PDF."
            ) from exc

    if pdfplumber_completed or extractor_failures == 0:
        raise ConverterNoTableError("Nenhuma tabela útil foi encontrada no PDF.")
    raise ConverterExtractionError("Os extratores de tabela não concluíram.")

# ============== Normalização de páginas (ATUALIZADO) ==============
def _papersize_token(name: str) -> str:
    n = (name or "A4").strip().lower()
    return n if n in ("a4", "letter") else "a4"

def _gs_autorotate_token(mode: str) -> str:
    mode = (mode or "none").strip().lower()
    if mode == "all":
        return "/All"
    if mode == "page":
        return "/PageByPage"
    return "/None"  # none

def normalize_pdf_pages(input_pdf: str, page_size: str = "A4", autorotate: str = "none") -> str:
    """
    Normaliza TODAS as páginas para A4/LETTER usando Ghostscript:
    -sPAPERSIZE=<a4|letter> -dFIXEDMEDIA -dPDFFitPage -dAutoRotatePages=<None|PageByPage|All>
    Retorna um **novo** caminho de saída.
    """
    token = _papersize_token(page_size)
    ar_token = _gs_autorotate_token(autorotate)
    root, ext = os.path.splitext(input_pdf)
    out_path = f"{root}_norm_{page_size.upper()}{ext or '.pdf'}"
    cmd = [
        GHOSTSCRIPT_BIN,
        "-sDEVICE=pdfwrite",
        f"-sPAPERSIZE={token}",
        "-dFIXEDMEDIA",
        "-dPDFFitPage",
        f"-dAutoRotatePages={ar_token}",
        "-dCompatibilityLevel=1.6",
        "-dNOPAUSE","-dBATCH","-dQUIET","-dSAFER",
        f"-sOutputFile={out_path}",
        input_pdf,
    ]
    configured_timeout = max(GHOSTSCRIPT_TIMEOUT, 60)
    effective_timeout = _effective_converter_timeout(
        configured_timeout,
        "ghostscript-normalize",
    )
    try:
        result = run_in_sandbox(
            cmd,
            cwd=os.path.dirname(os.path.abspath(input_pdf)),
            timeout=effective_timeout,
            cpu_seconds=max(1, math.ceil(effective_timeout)),
            mem_mb=768,
            output_limit_chars=4096,
        )
    except FileNotFoundError:
        raise ConverterToolUnavailableError(
            "Ghostscript não está disponível."
        ) from None
    except subprocess.TimeoutExpired:
        raise ConverterTimeoutError(
            "Ghostscript excedeu o tempo permitido."
        ) from None
    if result.returncode != 0:
        raise ConverterToolExecutionError(
            "Ghostscript terminou com erro."
        )
    check_converter_deadline("ghostscript-normalize-output")
    return out_path

def _strip_page_rotate(in_pdf: str) -> str:
    """
    Zera o /Rotate de todas as páginas (se existir), gerando um novo PDF.
    Útil quando PDFs trazem rotação fixa gravada.
    """
    fd, out_pdf = tempfile.mkstemp(
        prefix=".rotate-",
        suffix=".pdf",
        dir=os.path.dirname(os.path.abspath(in_pdf)),
    )
    os.close(fd)
    try:
        check_converter_deadline("merge-strip-rotate")
        try:
            from PyPDF2 import PdfReader, PdfWriter
        except Exception:
            from pypdf import PdfReader, PdfWriter
        reader = PdfReader(in_pdf)
        writer = PdfWriter()
        for pg in reader.pages:
            check_converter_deadline("merge-strip-rotate-page")
            # remove explicit rotate if present
            if "/Rotate" in pg:
                try:
                    # pypdf >= 3
                    pg.rotate(0)  # no-op, but keeps API similar
                    del pg["/Rotate"]
                except Exception:
                    try:
                        del pg["/Rotate"]
                    except Exception:
                        pass
            writer.add_page(pg)
        with open(out_pdf, "wb") as fh:
            writer.write(fh)
        check_converter_deadline("merge-strip-rotate-output")
        return out_pdf
    except ConverterTimeoutError:
        try:
            os.remove(out_pdf)
        except OSError:
            pass
        raise
    except Exception as e:
        logger.warning("Falha ao stripar /Rotate (%s). Retornando original.", e)
        try:
            os.remove(out_pdf)
        except Exception:
            pass
        return in_pdf

def _needs_normalization(in_pdf: str, page_size: str = "A4") -> bool:
    """
    Heurística: normalizar se:
      - houver páginas com tamanhos diferentes do alvo (A4/Letter), OU
      - houver mistura retrato/paisagem, OU
      - houver /Rotate explícito em qualquer página.
    """
    target = page_size.upper()
    tw, th = SIZES_PT.get(target, SIZES_PT["A4"])
    # tolerância em pontos
    tol = 2.0

    try:
        try:
            from PyPDF2 import PdfReader
        except Exception:
            from pypdf import PdfReader
        rdr = PdfReader(in_pdf)
        saw_portrait, saw_landscape = False, False
        for p in rdr.pages:
            mb = p.mediabox
            w = float(mb.right) - float(mb.left)
            h = float(mb.top) - float(mb.bottom)
            if abs(w - tw) > tol or abs(h - th) > tol:
                # aceita também A4 landscape (troca w/h)
                if not (abs(w - th) <= tol and abs(h - tw) <= tol):
                    return True
            if w >= h: saw_landscape = True
            else:      saw_portrait  = True
            # /Rotate explícito?
            try:
                if "/Rotate" in p and int(p["/Rotate"]) % 360 != 0:
                    return True
            except Exception:
                pass
        # se tiver mistura de orientações, ainda pode querer normalizar para A4 único
        return saw_portrait and saw_landscape
    except Exception as e:
        logger.debug("Falha ao inspecionar PDF (%s). Por segurança: normaliza.", e)
        return True

# ---------------- Dispatcher principal (mantido, agora usa o novo _image_to_pdf) ----------------
def _completed_converter_output(path: str) -> str:
    check_converter_deadline("converter-output")
    return path


def convert_upload_to_target(upload_file, target: str, out_dir: str) -> str:
    check_converter_deadline("converter-input")
    os.makedirs(out_dir, exist_ok=True)
    target = target.lower().strip()
    if target not in {'pdf','docx','csv','xlsm','xlsx'}:
        raise BadRequest(f"Destino não suportado: {target}")

    name = upload_file.filename or 'arquivo'
    ext = name.rsplit('.', 1)[-1].lower() if '.' in name else ''
    base = os.path.splitext(os.path.basename(name))[0] or 'arquivo'

    if target == 'pdf':
        if ext == 'pdf':
            in_path = _save_upload_to_tmp(
                upload_file,
                suffix='.pdf',
                directory=out_dir,
            )
            try:
                enforce_pdf_page_limit(in_path, label="PDF de entrada")
                dst_path = _unique_out_path(out_dir, base, 'pdf'); shutil.move(in_path, dst_path)
                return _completed_converter_output(dst_path)
            finally:
                try: os.remove(in_path)
                except OSError: pass

        if ext in IMG_EXTS:
            in_path = _save_upload_to_tmp(
                upload_file,
                suffix='.' + ext,
                directory=out_dir,
            )
            tmp_pdf = _unique_out_path(out_dir, base, "pdf")
            try:
                _image_to_pdf(in_path, tmp_pdf)   # ⬅️ A4/Letter + margens + auto-landscape + EXIF
                enforce_pdf_page_limit(tmp_pdf, label="PDF gerado")
                return _completed_converter_output(tmp_pdf)
            finally:
                try: os.remove(in_path)
                except OSError: pass

        if ext in DOC_EXTS or ext in SHEET_EXTS:
            in_path = _save_upload_to_tmp(
                upload_file,
                suffix='.' + ext if ext else '',
                directory=out_dir,
            )
            try:
                produced = _lo_convert(in_path, out_dir, 'pdf')
                enforce_pdf_page_limit(produced, label="PDF gerado")
                return _completed_converter_output(produced)
            finally:
                try: os.remove(in_path)
                except OSError: pass

        raise BadRequest(f'Extensão não suportada para conversão a PDF: {ext or "sem extensão"}')

    if target == 'docx':
        if ext in IMG_EXTS:
            raise BadRequest("Imagens não são convertidas para DOCX automaticamente.")
        in_path = _save_upload_to_tmp(
            upload_file,
            suffix='.' + ext if ext else '',
            directory=out_dir,
        )
        try:
            if ext == 'pdf':
                return _completed_converter_output(
                    _pdf_to_docx(in_path, out_dir)
                )
            return _completed_converter_output(
                _lo_convert(in_path, out_dir, 'docx', filter_name=FILTER_DOCX)
            )
        finally:
            try: os.remove(in_path)
            except OSError: pass

    if target == 'csv':
        in_path = _save_upload_to_tmp(
            upload_file,
            suffix='.' + ext if ext else '',
            directory=out_dir,
        )
        try:
            if ext == 'pdf':
                return _completed_converter_output(
                    _pdf_to_csv(in_path, out_dir)
                )
            if ext not in SHEET_EXTS:
                raise BadRequest("Apenas planilhas (xls/xlsx/ods/csv) ou PDF podem virar CSV.")
            return _completed_converter_output(
                _lo_convert(
                    in_path,
                    out_dir,
                    'csv',
                    filter_name=FILTER_CSV,
                    filter_opts=CSV_FILTER_OPTS,
                )
            )
        finally:
            try: os.remove(in_path)
            except OSError: pass

    if target == 'xlsm':
        in_path = _save_upload_to_tmp(
            upload_file,
            suffix='.' + ext if ext else '',
            directory=out_dir,
        )
        try:
            if ext == 'pdf':
                # Usa caminho de XLSX (rápido); caso necessário, aplique template XLSM fora daqui
                return _completed_converter_output(
                    _pdf_to_xlsx(in_path, out_dir)
                )
            if ext not in SHEET_EXTS:
                raise BadRequest("Apenas PDF ou planilhas (xls/xlsx/ods/csv) podem virar XLSM.")
            if ext == 'csv':
                return _completed_converter_output(
                    _convert_csv_to_spreadsheet(
                        in_path,
                        out_dir,
                        'xlsm',
                        FILTER_XLSM,
                    )
                )
            return _completed_converter_output(
                _lo_convert(in_path, out_dir, 'xlsm', filter_name=FILTER_XLSM)
            )
        finally:
            try: os.remove(in_path)
            except OSError: pass

    if target == 'xlsx':
        in_path = _save_upload_to_tmp(
            upload_file,
            suffix='.' + ext if ext else '',
            directory=out_dir,
        )
        try:
            if ext == 'pdf':
                return _completed_converter_output(
                    _pdf_to_xlsx(in_path, out_dir)
                )
            if ext == 'csv':
                return _completed_converter_output(
                    _convert_csv_to_spreadsheet(
                        in_path,
                        out_dir,
                        'xlsx',
                        FILTER_XLSX,
                    )
                )
            if ext in SHEET_EXTS:
                return _completed_converter_output(
                    _lo_convert(
                        in_path,
                        out_dir,
                        'xlsx',
                        filter_name=FILTER_XLSX,
                    )
                )
            raise BadRequest("Apenas PDF ou planilhas (xls/xlsx/ods/csv) podem virar XLSX.")
        finally:
            try: os.remove(in_path)
            except OSError: pass

    raise BadRequest(f"Destino não suportado: {target}")

# ---------------- Legacy compat (usa o novo _image_to_pdf) ----------------
def converter_doc_para_pdf(upload_file, modificacoes=None) -> str:
    """Compat antigo: converte qualquer documento/imagem para PDF."""
    name = upload_file.filename or 'arquivo'
    ext = name.rsplit('.', 1)[-1].lower() if '.' in name else ''
    in_path = _save_upload_to_tmp(upload_file, suffix='.' + ext if ext else '')
    out_path = _tmp_out_path('pdf')
    try:
        if ext == 'pdf':
            shutil.move(in_path, out_path)
            enforce_pdf_page_limit(out_path, label="PDF de entrada")
            return out_path
        if ext in IMG_EXTS:
            _image_to_pdf(in_path, out_path)  # ⬅️ A4/Letter + EXIF agora
        elif ext in DOC_EXTS or ext in SHEET_EXTS:
            produced = _lo_convert(in_path, os.path.dirname(out_path), 'pdf')
            if os.path.abspath(produced) != os.path.abspath(out_path):
                if os.path.exists(out_path):
                    os.remove(out_path)
                shutil.move(produced, out_path)
        else:
            raise BadRequest(f'Extensão não suportada para este conversor: {ext or "sem extensão"}')
        enforce_pdf_page_limit(out_path, label="PDF gerado")
        return out_path
    finally:
        try: os.remove(in_path)
        except OSError: pass

def converter_planilha_para_pdf(upload_file, modificacoes=None) -> str:
    """Compat antigo: planilhas (xls/xlsx/ods/csv) → PDF via LibreOffice."""
    name = upload_file.filename or 'planilha'
    ext = name.rsplit('.', 1)[-1].lower() if '.' in name else ''
    in_path = _save_upload_to_tmp(upload_file, suffix='.' + ext if ext else '')
    out_path = _tmp_out_path('pdf')
    try:
        produced = _lo_convert(in_path, os.path.dirname(out_path), 'pdf')
        if os.path.abspath(produced) != os.path.abspath(out_path):
            if os.path.exists(out_path):
                os.remove(out_path)
            shutil.move(produced, out_path)
        enforce_pdf_page_limit(out_path, label="PDF gerado")
        return out_path
    finally:
        try: os.remove(in_path)
        except OSError: pass

# ---------------- Multi-file (compat) ----------------
def convert_many_uploads(files, target: str, out_dir: str):
    outputs = []
    os.makedirs(out_dir, exist_ok=True)
    for up in files:
        check_converter_deadline("converter-next-file")
        outputs.append(convert_upload_to_target(up, target, out_dir))
    return outputs

# --- Unir vários uploads em UM PDF (normaliza A4 por padrão) ---
def convert_many_uploads_to_single_pdf(
    uploads: List,
    workdir: str | None = None,
    *,
    normalize: str = None,  # "auto"|"on"|"off" ; se None, usa MERGE_NORMALIZE_MODE
    norm_page_size: str = "A4",
) -> str:
    check_converter_deadline("merge-input")
    if not uploads:
        raise ValueError("Nenhum arquivo enviado.")

    try:
        from PyPDF2 import PdfMerger  # pip install PyPDF2
    except Exception:
        from pypdf import PdfMerger   # pip install pypdf

    out_dir = os.path.abspath(workdir) if workdir else tempfile.mkdtemp(prefix="gvpdf_merge_")
    os.makedirs(out_dir, exist_ok=True)

    pdf_paths = convert_many_uploads(uploads, 'pdf', out_dir)
    check_converter_deadline("merge-conversions")
    pdf_paths = [p for p in (pdf_paths or []) if p and os.path.isfile(p)]
    if not pdf_paths:
        raise RuntimeError("Conversão não gerou PDFs.")

    final_path = _unique_out_path(out_dir, "arquivos_unidos", "pdf")
    merger = PdfMerger()
    try:
        for p in pdf_paths:
            check_converter_deadline("merge-append")
            merger.append(p)
        with open(final_path, "wb") as fh:
            merger.write(fh)
    finally:
        merger.close()

    # (opcional) remove /Rotate das páginas antes de normalizar
    merged_path = final_path
    if MERGE_STRIP_ROTATE:
        check_converter_deadline("merge-strip-rotate")
        stripped = _strip_page_rotate(final_path)
        if stripped != final_path:
            try: os.remove(final_path)
            except OSError: pass
            merged_path = stripped

    # Decide normalização
    norm_mode = (normalize or MERGE_NORMALIZE_MODE or "auto").lower()
    if norm_mode != "off":
        check_converter_deadline("merge-normalize")
        try:
            if norm_mode == "always" or (norm_mode == "auto" and _needs_normalization(merged_path, norm_page_size)):
                normalized = normalize_pdf_pages(merged_path, norm_page_size, autorotate=MERGE_NORMALIZE_AUTOROTATE)
                try: os.remove(merged_path)
                except OSError: pass
                merged_path = normalized
        except ConverterTimeoutError:
            raise
        except Exception as e:
            logger.warning("Normalização falhou (%s); retornando merge bruto.", e)
    check_converter_deadline("merge-output")

    # ===== (7.1) MÉTRICAS: registrar 'convert' (multi-upload → 1 PDF) =====
    try:
        # bytes_in ~ soma dos PDFs gerados a partir dos uploads
        bytes_in = 0
        for p in pdf_paths:
            try:
                bytes_in += os.path.getsize(p)
            except Exception:
                pass
        # se houver contexto de request e Content-Length maior, usar como melhor estimativa
        if has_request_context():
            try:
                cl = int(request.content_length or 0)
                if cl > bytes_in:
                    bytes_in = cl
            except Exception:
                pass
        bytes_out = os.path.getsize(merged_path) if os.path.exists(merged_path) else None
        record_job_event(
            route="/api/convert/merge-a4",  # chama /convert/merge-a4 (e /to-pdf-merge alias)
            action="convert",
            bytes_in=(bytes_in if bytes_in > 0 else None),
            bytes_out=bytes_out,
            files_out=1,
        )
    except Exception:
        pass
    # =====================================================================

    return merged_path

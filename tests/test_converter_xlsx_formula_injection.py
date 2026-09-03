import ast
import inspect
import logging
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from app.services import converter_service
from app.services.converter_output_validation import validate_converter_output


DANGEROUS_TEXTS = [
    "=1+1",
    '=HYPERLINK("https://example.invalid";"abrir")',
    "+SUM(A1:A2)",
    "-10+20",
    "@SUM(A1:A2)",
    "\t=1+1",
    "\r=1+1",
    "\n=1+1",
    "   =1+1",
    "  \t=1+1",
    "=cmd|' /C calc'!A0",
    "   -10",
]

CONTROL_PREFIX_TEXTS = [
    "\x00=1+1",
    "\x1b@SUM(A1:A2)",
    "\u200b+1",
]


def _write_dataframe(path: Path, dataframe: pd.DataFrame) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        converter_service._write_dataframe_to_spreadsheet(
            dataframe,
            writer,
            sheet_name="Dados",
        )


@pytest.mark.parametrize("payload", DANGEROUS_TEXTS + CONTROL_PREFIX_TEXTS)
def test_dangerous_text_is_neutralized_idempotently(payload):
    neutralized = converter_service._neutralize_spreadsheet_formula(payload)

    assert neutralized == "'" + payload
    assert (
        converter_service._neutralize_spreadsheet_formula(neutralized)
        == neutralized
    )


@pytest.mark.parametrize(
    "value",
    [
        "texto seguro",
        "10",
        "  texto seguro",
        "'=já neutralizado",
        "   '=já neutralizado",
        "",
        "   ",
    ],
)
def test_safe_text_is_preserved_exactly(value):
    assert converter_service._neutralize_spreadsheet_formula(value) == value


@pytest.mark.parametrize(
    "value",
    [
        None,
        -10,
        -10.5,
        Decimal("-10"),
        True,
        False,
        date(2026, 7, 27),
        datetime(2026, 7, 27, 12, 30),
        time(12, 30),
    ],
)
def test_non_text_values_preserve_value_and_type(value):
    result = converter_service._neutralize_spreadsheet_formula(value)

    assert result == value
    assert type(result) is type(value)


def test_dataframe_barrier_preserves_original_shape_labels_types_and_duplicates():
    original = pd.DataFrame(
        {
            "=Cabeçalho": ["=1+1", "=1+1"],
            "Número": [-10, -10],
            "Decimal": [Decimal("-2.5"), Decimal("-2.5")],
            "Data": [date(2026, 7, 27), date(2026, 7, 27)],
            "Flag": [True, True],
            "Nulo": [None, None],
            "Vazio": ["", ""],
        },
        index=pd.Index([7, 7], name="origem"),
    )
    snapshot = original.copy(deep=True)

    protected = converter_service._neutralize_dataframe_for_spreadsheet(
        original
    )

    pd.testing.assert_frame_equal(original, snapshot)
    assert protected is not original
    assert protected.shape == original.shape
    assert protected.index.equals(original.index)
    assert protected.index.name == original.index.name
    assert list(protected.columns) == list(original.columns)
    assert protected.duplicated().sum() == original.duplicated().sum()
    assert protected["=Cabeçalho"].tolist() == ["'=1+1", "'=1+1"]
    assert protected["Número"].tolist() == [-10, -10]
    assert all(
        isinstance(value, Decimal)
        for value in protected["Decimal"]
    )
    assert all(isinstance(value, date) for value in protected["Data"])
    assert protected["Flag"].tolist() == [True, True]
    assert protected["Nulo"].isna().all()
    assert protected["Vazio"].tolist() == ["", ""]


def test_generic_writer_serializes_payloads_and_header_as_literal_text(tmp_path):
    output = tmp_path / "generic.xlsx"
    source = pd.DataFrame(
        {
            "=Cabeçalho": DANGEROUS_TEXTS,
            "Número": [-10] * len(DANGEROUS_TEXTS),
            "Data": [date(2026, 7, 27)] * len(DANGEROUS_TEXTS),
            "Flag": [True] * len(DANGEROUS_TEXTS),
        }
    )
    snapshot = source.copy(deep=True)

    _write_dataframe(output, source)

    pd.testing.assert_frame_equal(source, snapshot)
    workbook = load_workbook(output, data_only=False)
    worksheet = workbook["Dados"]
    try:
        assert worksheet.max_row == len(source) + 1
        assert worksheet.max_column == len(source.columns)
        assert worksheet["A1"].value == "'=Cabeçalho"
        assert worksheet["A1"].data_type != "f"
        for row_index, payload in enumerate(DANGEROUS_TEXTS, start=2):
            cell = worksheet.cell(row=row_index, column=1)
            assert cell.value == "'" + payload
            assert cell.data_type != "f"
            assert worksheet.cell(row=row_index, column=2).value == -10
            assert worksheet.cell(row=row_index, column=2).data_type == "n"
            assert worksheet.cell(row=row_index, column=3).data_type == "d"
            assert worksheet.cell(row=row_index, column=4).value is True
    finally:
        workbook.close()

    calculated_view = load_workbook(output, data_only=True)
    try:
        assert calculated_view["Dados"]["A2"].value == "'" + DANGEROUS_TEXTS[0]
    finally:
        calculated_view.close()

    assert (
        validate_converter_output(str(output), str(tmp_path), "xlsx")
        == str(output.resolve())
    )


def test_minimal_xlsx_writer_uses_the_same_literal_barrier(tmp_path):
    output = tmp_path / "minimal.xlsx"
    payload = '=HYPERLINK("https://example.invalid";"abrir")'

    converter_service._write_minimal_xlsx(str(output), payload)

    workbook = load_workbook(output, data_only=False)
    try:
        cell = workbook["Dados"]["A2"]
        assert cell.value == "'" + payload
        assert cell.data_type != "f"
    finally:
        workbook.close()


def test_coparticipation_normalizer_uses_same_literal_writer_barrier(
    tmp_path,
    caplog,
):
    payload = '=HYPERLINK("https://example.invalid";"abrir")'
    raw = pd.DataFrame([
        ["1.01.01012", payload, "R$ 100,00", "R$ 20,00"],
    ])
    normalized = converter_service._normalize_coparticipacao_table(raw)
    expected_columns = [
        "Secao",
        "Codigo",
        "Procedimento",
        "Valor Unimed",
        "Copart 20%",
        "Copart 30%",
        "Copart 40%",
        "Copart 50%",
        "Tipo Linha",
    ]
    output = tmp_path / "coparticipacao.xlsx"

    assert list(normalized.columns) == expected_columns
    assert normalized.loc[0, "Procedimento"] == payload

    with caplog.at_level(logging.DEBUG):
        _write_dataframe(output, normalized)

    workbook = load_workbook(output, data_only=False)
    worksheet = workbook["Dados"]
    try:
        assert [
            worksheet.cell(row=1, column=index).value
            for index in range(1, len(expected_columns) + 1)
        ] == expected_columns
        procedure_cell = worksheet["C2"]
        assert procedure_cell.value == "'" + payload
        assert procedure_cell.data_type != "f"
        assert worksheet.max_column == len(expected_columns)
    finally:
        workbook.close()

    assert payload not in caplog.text
    assert (
        validate_converter_output(str(output), str(tmp_path), "xlsx")
        == str(output.resolve())
    )


def test_all_application_dataframe_writes_cross_the_single_barrier():
    app_root = Path(converter_service.__file__).resolve().parents[1]
    sinks = []

    for python_path in app_root.rglob("*.py"):
        source = python_path.read_text(encoding="utf-8")
        tree = ast.parse(source)
        relative_path = python_path.relative_to(app_root).as_posix()

        class ToExcelVisitor(ast.NodeVisitor):
            def __init__(self):
                self.functions = []

            def visit_FunctionDef(self, node):
                self.functions.append(node.name)
                self.generic_visit(node)
                self.functions.pop()

            def visit_Call(self, node):
                if (
                    isinstance(node.func, ast.Attribute)
                    and node.func.attr == "to_excel"
                ):
                    function = self.functions[-1] if self.functions else None
                    sinks.append((relative_path, function))
                self.generic_visit(node)

        ToExcelVisitor().visit(tree)

    assert sinks == [
        ("services/converter_service.py", "_write_dataframe_to_spreadsheet")
    ]
    assert "_write_dataframe_to_spreadsheet(" in inspect.getsource(
        converter_service._pdf_to_xlsx
    )
    assert "_neutralize_spreadsheet_formula" in inspect.getsource(
        converter_service._write_minimal_xlsx
    )


def test_pdf_csv_and_candidate_pipeline_do_not_use_spreadsheet_barrier():
    unaffected_functions = [
        converter_service._pdf_to_csv,
        converter_service._make_table_candidate,
        converter_service._score_table_candidate,
        converter_service._select_and_deduplicate_candidates,
        converter_service._map_pdf_table_geometry_shadow,
    ]

    for function in unaffected_functions:
        source = inspect.getsource(function)
        assert "_neutralize_dataframe_for_spreadsheet" not in source
        assert "_write_dataframe_to_spreadsheet" not in source


def test_xlsm_route_does_not_share_pdf_xlsx_dataframe_writer():
    source = inspect.getsource(converter_service.convert_upload_to_target)
    xlsm_branch = source.split("if target == 'xlsm':", 1)[1].split(
        "if target == 'xlsx':",
        1,
    )[0]

    assert "_lo_convert" in xlsm_branch
    assert "_write_dataframe_to_spreadsheet" not in xlsm_branch

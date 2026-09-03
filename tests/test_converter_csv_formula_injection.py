import csv
import io
import inspect
import logging
import os
import subprocess
import zipfile
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree

import pytest
from openpyxl import Workbook, load_workbook
from werkzeug.datastructures import FileStorage

from app import create_app
from app.services import converter_service
from app.services.converter_output_validation import validate_converter_output


DANGEROUS_FIELDS = [
    "=1+1",
    "+SUM(A1:A2)",
    "-10+20",
    "@SUM(A1:A2)",
    "   =1+1",
    "\t=1+1",
    "\r=1+1",
    "\n=1+1",
    '=HYPERLINK("https://example.invalid";"open")',
    "='file:///C:/tmp/does-not-exist.ods'#$Sheet1.A1",
    '=DDE("synthetic.invalid";"topic";"item")',
]


def _csv_bytes(rows, *, delimiter=","):
    stream = io.StringIO(newline="")
    writer = csv.writer(
        stream,
        delimiter=delimiter,
        quoting=csv.QUOTE_ALL,
        lineterminator="\n",
    )
    for row in rows:
        writer.writerow(row)
    return ("\ufeff" + stream.getvalue()).encode("utf-8")


def _read_csv(path, *, delimiter):
    with open(path, "r", encoding="utf-8-sig", newline="") as stream:
        return list(csv.reader(stream, delimiter=delimiter, strict=True))


def _formula_elements(path):
    formulas = []
    with zipfile.ZipFile(path, "r") as archive:
        for member in archive.namelist():
            if not (
                member.startswith("xl/worksheets/")
                and member.endswith(".xml")
            ):
                continue
            root = ElementTree.fromstring(archive.read(member))
            for element in root.iter():
                if element.tag.rsplit("}", 1)[-1] == "f":
                    formulas.append((member, element.text or ""))
    return formulas


def _write_workbook(path, *, formula=None):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = formula if formula is not None else "texto seguro"
    workbook.save(path)
    workbook.close()


@pytest.mark.parametrize("value", DANGEROUS_FIELDS)
def test_csv_dangerous_field_uses_central_idempotent_policy(value):
    protected = converter_service._neutralize_csv_field_for_spreadsheet(value)

    assert protected == "'" + value
    assert (
        converter_service._neutralize_csv_field_for_spreadsheet(protected)
        == protected
    )


@pytest.mark.parametrize(
    "value",
    [
        "texto seguro",
        "",
        "   ",
        "'=já neutralizado",
    ],
)
def test_csv_safe_empty_whitespace_and_preprotected_fields_are_unchanged(value):
    assert converter_service._neutralize_csv_field_for_spreadsheet(value) == value


@pytest.mark.parametrize(
    "value",
    [
        "-10",
        "-10.5",
        "-10,5",
        "-1E2",
        "+10",
        "  -10  ",
    ],
)
def test_complete_unambiguous_numeric_tokens_are_not_neutralized(value):
    assert converter_service._neutralize_csv_field_for_spreadsheet(value) == value


@pytest.mark.parametrize(
    "value",
    [
        "-10+20",
        "+1+1",
        "-1E2+3",
        "-10 trailing",
    ],
)
def test_numeric_prefix_with_remaining_content_is_neutralized(value):
    assert (
        converter_service._neutralize_csv_field_for_spreadsheet(value)
        == "'" + value
    )


@pytest.mark.parametrize("delimiter", [",", ";"])
def test_neutralized_csv_copy_preserves_dialect_structure_and_content(
    tmp_path,
    delimiter,
):
    embedded_delimiter = f"parte{delimiter}interna"
    embedded_newline = "linha 1\nlinha 2"
    rows = [
        ["cabecalho", "valor"],
        ["perigoso", "=1+1"],
        ["delimitador", embedded_delimiter],
        ["multilinha", embedded_newline],
        [],
        ["duplicado", "mesmo"],
        ["duplicado", "mesmo"],
        ["numero", "-10"],
        ["seguro", "texto seguro"],
    ]
    source = tmp_path / "entrada.csv"
    source.write_bytes(_csv_bytes(rows, delimiter=delimiter))
    original_bytes = source.read_bytes()

    protected, detected_delimiter = (
        converter_service._write_neutralized_csv_copy(
            str(source),
            str(tmp_path),
        )
    )

    try:
        assert detected_delimiter == delimiter
        assert Path(protected) != source
        assert source.read_bytes() == original_bytes
        parsed = _read_csv(protected, delimiter=delimiter)
        expected = [list(row) for row in rows]
        expected[1][1] = "'=1+1"
        assert parsed == expected
        assert parsed[2][1] == embedded_delimiter
        assert parsed[3][1] == embedded_newline
        assert parsed[4] == []
        assert parsed[5] == parsed[6]
    finally:
        os.remove(protected)


def test_ooxml_formula_guard_accepts_literal_cells_and_rejects_formula(tmp_path):
    safe = tmp_path / "safe.xlsx"
    unsafe = tmp_path / "unsafe.xlsx"
    _write_workbook(safe)
    _write_workbook(unsafe, formula="=1+1")

    converter_service._assert_spreadsheet_has_no_formulas(str(safe))
    with pytest.raises(converter_service.ConverterToolExecutionError):
        converter_service._assert_spreadsheet_has_no_formulas(str(unsafe))


@pytest.mark.parametrize("failure", ["success", "error", "timeout"])
def test_csv_copy_cleanup_on_success_error_and_timeout(
    tmp_path,
    monkeypatch,
    failure,
):
    source = tmp_path / "entrada.csv"
    source.write_bytes(_csv_bytes([["=1+1"]]))
    converted_inputs = []

    def fake_lo_convert(
        input_path,
        out_dir,
        out_ext,
        filter_name=None,
        **kwargs,
    ):
        converted_inputs.append(Path(input_path))
        assert kwargs['input_filter'] == (
            f'{converter_service.FILTER_CSV}:'
            '44,34,76,1,,1033,false,false'
        )
        assert Path(input_path) != source
        assert Path(input_path).exists()
        if failure == "error":
            raise converter_service.ConverterToolExecutionError(
                "falha sintética"
            )
        if failure == "timeout":
            raise subprocess.TimeoutExpired(["soffice"], 1)
        output = Path(out_dir) / f"resultado.{out_ext}"
        _write_workbook(output)
        return str(output)

    monkeypatch.setattr(converter_service, "_lo_convert", fake_lo_convert)

    if failure == "success":
        result = converter_service._convert_csv_to_spreadsheet(
            str(source),
            str(tmp_path),
            "xlsx",
            converter_service.FILTER_XLSX,
        )
        assert Path(result).exists()
    else:
        with pytest.raises(
            (
                converter_service.ConverterToolExecutionError
                if failure == "error"
                else subprocess.TimeoutExpired
            )
        ):
            converter_service._convert_csv_to_spreadsheet(
                str(source),
                str(tmp_path),
                "xlsx",
                converter_service.FILTER_XLSX,
            )

    assert converted_inputs
    assert all(not path.exists() for path in converted_inputs)
    assert source.exists()


def test_residual_formula_removes_output_and_neutralized_copy(
    tmp_path,
    monkeypatch,
):
    source = tmp_path / "entrada.csv"
    source.write_bytes(_csv_bytes([["=1+1"]]))
    captured = {}

    def fake_lo_convert(
        input_path,
        out_dir,
        out_ext,
        filter_name=None,
        **kwargs,
    ):
        captured["protected"] = Path(input_path)
        output = Path(out_dir) / f"resultado.{out_ext}"
        _write_workbook(output, formula="=1+1")
        captured["output"] = output
        return str(output)

    monkeypatch.setattr(converter_service, "_lo_convert", fake_lo_convert)

    with pytest.raises(converter_service.ConverterToolExecutionError):
        converter_service._convert_csv_to_spreadsheet(
            str(source),
            str(tmp_path),
            "xlsx",
            converter_service.FILTER_XLSX,
        )

    assert not captured["protected"].exists()
    assert not captured["output"].exists()
    assert source.exists()


@pytest.mark.parametrize("source_ext", ["xls", "xlsx", "ods"])
@pytest.mark.parametrize("target", ["xlsx", "xlsm"])
def test_non_csv_spreadsheets_bypass_csv_neutralization(
    tmp_path,
    monkeypatch,
    source_ext,
    target,
):
    calls = []

    def forbidden_csv_copy(*_args, **_kwargs):
        raise AssertionError("neutralização CSV não deveria ser usada")

    def fake_lo_convert(
        input_path,
        out_dir,
        out_ext,
        filter_name=None,
        **kwargs,
    ):
        calls.append((Path(input_path).suffix, out_ext, filter_name))
        output = Path(out_dir) / f"resultado.{out_ext}"
        output.write_bytes(b"synthetic workbook")
        return str(output)

    monkeypatch.setattr(
        converter_service,
        "_write_neutralized_csv_copy",
        forbidden_csv_copy,
    )
    monkeypatch.setattr(converter_service, "_lo_convert", fake_lo_convert)
    upload = FileStorage(
        stream=BytesIO(b"synthetic workbook"),
        filename=f"entrada.{source_ext}",
    )

    output = converter_service.convert_upload_to_target(
        upload,
        target,
        str(tmp_path),
    )

    assert Path(output).exists()
    assert calls == [
        (
            f".{source_ext}",
            target,
            (
                converter_service.FILTER_XLSX
                if target == "xlsx"
                else converter_service.FILTER_XLSM
            ),
        )
    ]


@pytest.mark.parametrize(
    ("target", "delimiter"),
    [("xlsx", ","), ("xlsm", ";")],
)
def test_real_csv_endpoint_produces_formula_free_structural_workbook(
    tmp_path,
    caplog,
    target,
    delimiter,
):
    rows = [["valor"]]
    rows.extend([[value] for value in DANGEROUS_FIELDS])
    rows.extend([
        ["texto seguro"],
        [""],
        ["   "],
        ["'=já neutralizado"],
        ["-10"],
        ["-10.5"],
        [f"campo{delimiter}citado"],
        ["linha 1\nlinha 2"],
        [],
        ["duplicado"],
        ["duplicado"],
    ])
    for row_index, row in enumerate(rows):
        if row:
            row.append(
                'duplicate-metadata'
                if row[0] == 'duplicado'
                else f'metadata-{row_index}'
            )
    app = create_app()
    app.config.update(
        TESTING=True,
        WTF_CSRF_ENABLED=False,
        RATELIMIT_ENABLED=False,
        UPLOAD_FOLDER=str(tmp_path),
        CONVERTER_MAX_FILES=2,
        CONVERTER_MAX_RUNTIME_SEC=120,
    )
    endpoint = f"/api/convert/to-{target}"
    caplog.set_level(logging.DEBUG)
    caplog.clear()

    client = app.test_client()
    response = client.post(
        endpoint,
        data={
            "files[]": (
                BytesIO(_csv_bytes(rows, delimiter=delimiter)),
                "synthetic.csv",
            )
        },
        content_type="multipart/form-data",
        headers={"Accept": "application/json"},
    )

    assert response.status_code == 200, response.get_data(as_text=True)
    response_payload = response.get_json()
    assert response_payload["count"] == 1
    assert set(response_payload["files"][0]) == {
        "name",
        "size",
        "download_url",
    }
    assert response_payload["files"][0]["name"] == f"synthetic.{target}"
    matches = list(tmp_path.rglob(f"synthetic.{target}"))
    download_url = response_payload['files'][0]['download_url']
    assert download_url.startswith('/viewer/raw/generated/')
    assert client.get(download_url).status_code == 200
    assert len(matches) == 1
    output = matches[0]
    assert (
        validate_converter_output(str(output), str(output.parent), target)
        == str(output.resolve())
    )

    workbook = load_workbook(
        output,
        data_only=False,
        keep_vba=(target == "xlsm"),
    )
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        values = [
            worksheet.cell(row=index, column=1).value
            for index in range(1, len(rows) + 1)
        ]
        assert worksheet.max_row == len(rows)
        assert worksheet.max_column == 2
        assert all(
            cell.data_type != "f"
            for row in worksheet.iter_rows()
            for cell in row
        )
        for row_index, original in enumerate(DANGEROUS_FIELDS, start=2):
            normalized_original = original.replace('\r\n', '\n').replace(
                '\r',
                '\n',
            )
            assert values[row_index - 1] == chr(39) + normalized_original
        assert values[len(DANGEROUS_FIELDS) + 1] == "texto seguro"
        for row_index, original_row in enumerate(rows, start=1):
            expected_metadata = (
                original_row[1] if len(original_row) > 1 else None
            )
            assert worksheet.cell(row_index, 2).value == expected_metadata
        assert -10 in values
        assert -10.5 in values
        assert f"campo{delimiter}citado" in values
        assert "linha 1\nlinha 2" in values
        assert values[-1] == values[-2] == "duplicado"
    finally:
        workbook.close()
        vba_archive = getattr(workbook, 'vba_archive', None)
        if vba_archive is not None:
            vba_archive.close()

    assert _formula_elements(output) == []
    if target == "xlsm":
        with zipfile.ZipFile(output, "r") as archive:
            assert (
                b"application/vnd.ms-excel.sheet.macroEnabled.main+xml"
                in archive.read("[Content_Types].xml")
            )
    for value in DANGEROUS_FIELDS:
        assert value not in caplog.text


@pytest.mark.parametrize('target', ['xlsx', 'xlsm'])
def test_residual_formula_is_not_published_and_logs_are_redacted(
    tmp_path,
    monkeypatch,
    caplog,
    target,
):
    formula = "=1+1"
    output_paths = []

    def fake_lo_convert(
        input_path,
        out_dir,
        out_ext,
        filter_name=None,
        **kwargs,
    ):
        output = Path(out_dir) / f"resultado.{out_ext}"
        _write_workbook(output, formula=formula)
        output_paths.append(output)
        return str(output)

    monkeypatch.setattr(converter_service, "_lo_convert", fake_lo_convert)
    app = create_app()
    app.config.update(
        TESTING=True,
        WTF_CSRF_ENABLED=False,
        RATELIMIT_ENABLED=False,
        UPLOAD_FOLDER=str(tmp_path),
    )
    caplog.set_level(logging.WARNING)
    caplog.clear()

    response = app.test_client().post(
        f'/api/convert/to-{target}',
        data={
            "files[]": (
                BytesIO(_csv_bytes([[formula]])),
                "synthetic.csv",
            )
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 503
    assert response.get_json() == {
        "error": "A ferramenta de conversão não concluiu o processamento."
    }
    assert not list(tmp_path.rglob(f'*.{target}'))
    assert output_paths and all(not path.exists() for path in output_paths)
    assert formula not in caplog.text


def test_pdf_paths_and_existing_xlsx_barrier_remain_separate():
    assert "_convert_csv_to_spreadsheet" not in inspect.getsource(
        converter_service._pdf_to_xlsx
    )
    assert "_convert_csv_to_spreadsheet" not in inspect.getsource(
        converter_service._pdf_to_csv
    )
    assert "_write_dataframe_to_spreadsheet" in inspect.getsource(
        converter_service._pdf_to_xlsx
    )
